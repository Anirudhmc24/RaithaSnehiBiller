import io
import datetime
import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter

from config.settings import SHOP_NAME, SHOP_ADDR1, SHOP_ADDR2, SHOP_PHONE, SHOP_STATE, GSTIN, MON_ABBR
from database.db_main import get_conn

def generate_excel_report() -> bytes:
    conn   = get_conn()
    today  = datetime.date.today()
    buffer = io.BytesIO()
    wb     = xlsxwriter.Workbook(buffer, {"in_memory": True})

    hdr_fmt   = wb.add_format({"bold":True,"bg_color":"#1a3c5e","font_color":"white",
                                "border":1,"align":"center","valign":"vcenter","text_wrap":True})
    cell_fmt  = wb.add_format({"border":1,"align":"left"})
    num_fmt   = wb.add_format({"border":1,"num_format":"#,##0.00","align":"right"})
    title_fmt = wb.add_format({"bold":True,"font_size":14,"font_color":"#1a3c5e"})
    sub_fmt   = wb.add_format({"italic":True,"font_color":"#555555"})
    total_fmt = wb.add_format({"bold":True,"bg_color":"#e8f5e9","border":1,
                                "num_format":"#,##0.00","align":"right"})

    COLS = [
        ("Date",15),("Invoice No",18),("Customer",25),("GSTIN",20),
        ("HSN",12),("Product",30),("GST Rate",10),("Qty",8),("Unit",8),
        ("Taxable Value",15),("CGST",13),("SGST",13),("Total",15),
    ]

    def write_sheet(name, where_sql, params):
        ws = wb.add_worksheet(name)
        ws.merge_range("A1:M1", f"{SHOP_NAME} — {name} Sales Report", title_fmt)
        ws.merge_range("A2:M2", f"GSTIN: {GSTIN}  |  Generated: {today:%d %b %Y}  |  For GSTR-1 Filing", sub_fmt)
        ws.set_row(0, 22)
        ws.set_row(2, 30)
        for ci, (cn, cw) in enumerate(COLS):
            ws.write(2, ci, cn, hdr_fmt)
            ws.set_column(ci, ci, cw)

        rows = conn.execute(f"""
            SELECT i.invoice_date, i.invoice_no, i.customer_name, i.customer_gstin,
                   ii.hsn_code, ii.product_name, ii.gst_rate, ii.quantity, ii.unit,
                   ii.taxable_value, ii.cgst_amount, ii.sgst_amount, ii.line_total
            FROM invoices i JOIN invoice_items ii ON i.invoice_no=ii.invoice_no
            WHERE {where_sql}
            ORDER BY i.invoice_date, i.invoice_no
        """, params).fetchall()

        ri     = 3
        totals = {"taxable_value": 0.0, "cgst_amount": 0.0, "sgst_amount": 0.0, "line_total":  0.0}
        for r in rows:
            ws.write(ri, 0,  r["invoice_date"],                          cell_fmt)
            ws.write(ri, 1,  r["invoice_no"],                            cell_fmt)
            ws.write(ri, 2,  r["customer_name"]  or "Cash Customer",     cell_fmt)
            ws.write(ri, 3,  r["customer_gstin"] or "Unregistered",      cell_fmt)
            ws.write(ri, 4,  r["hsn_code"],                              cell_fmt)
            ws.write(ri, 5,  r["product_name"],                          cell_fmt)
            ws.write(ri, 6,  f"{(r['gst_rate'] or 0)*100:.0f}%",        cell_fmt)
            ws.write(ri, 7,  r["quantity"],                              num_fmt)
            ws.write(ri, 8,  r["unit"],                                  cell_fmt)
            ws.write(ri, 9,  r["taxable_value"],                         num_fmt)
            ws.write(ri, 10, r["cgst_amount"],                           num_fmt)
            ws.write(ri, 11, r["sgst_amount"],                           num_fmt)
            ws.write(ri, 12, r["line_total"],                            num_fmt)
            
            totals["taxable_value"] += r["taxable_value"] or 0
            totals["cgst_amount"]   += r["cgst_amount"]   or 0
            totals["sgst_amount"]   += r["sgst_amount"]   or 0
            totals["line_total"]    += r["line_total"]    or 0
            ri += 1

        if rows:
            ws.merge_range(ri, 0, ri, 8, "TOTAL", hdr_fmt)
            ws.write(ri, 9,  totals["taxable_value"], total_fmt)
            ws.write(ri, 10, totals["cgst_amount"],   total_fmt)
            ws.write(ri, 11, totals["sgst_amount"],   total_fmt)
            ws.write(ri, 12, totals["line_total"],    total_fmt)
        else:
            ws.merge_range(ri, 0, ri, 12, "No transactions for this period.", cell_fmt)
        ws.freeze_panes(3, 0)

    today_s    = today.isoformat()
    week_start = today - datetime.timedelta(days=today.weekday())
    week_end   = week_start + datetime.timedelta(days=6)

    write_sheet("Daily",   "DATE(i.invoice_date)=?",               (today_s,))
    write_sheet("Weekly",  "DATE(i.invoice_date) BETWEEN ? AND ?", (week_start.isoformat(), week_end.isoformat()))
    write_sheet("Monthly", "strftime('%Y-%m',i.invoice_date)=?",   (today.strftime("%Y-%m"),))

    conn.close()
    wb.close()
    return buffer.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# GSTR-1 Sales & Purchase Registers (XLSX)
# ─────────────────────────────────────────────────────────────────────────────
from utils.helpers import parse_mk, ld, r2

def header_block(reg_type, m, y):
    ys=str(y)[2:]; yn=str(y+1)[2:]; abbr=MON_ABBR[m]; last=ld(m,y)
    return [[f"{SHOP_NAME}-{ys}-{yn}"],[SHOP_ADDR1],[SHOP_ADDR2],
            [SHOP_PHONE],[SHOP_STATE],[reg_type],
            [f"1-{abbr}-{ys} to {last}-{abbr}-{ys}"]]

def set_widths(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

def to_bytes(wb):
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

def split_evenly(total, n):
    if not total or n<1: return [0.0]*max(n,1)
    base=r2(total/n); parts=[base]*n; parts[-1]=r2(total-base*(n-1))
    return parts

def make_sales_xlsx(mk, entries, start_vno):
    m,y=parse_mk(mk)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Sales Register"
    for row in header_block("Sales Register",m,y): ws.append(row)
    ws.append(["Date","Particulars","Buyer","Voucher Type","Voucher No.",
        "Voucher Ref. No.","GSTIN/UIN","Value","Gross Total",
        "18% Sales","9% CGST Output","9% SGST  Output",
        "5% Sales","2.5% Cgst Output","2.5% Sgst Output",
        "12% Sales","6% CGST Output","6% SGCT Output","Exampted Sales"])
    vno=start_vno
    g={k:0.0 for k in ["val","gross","v18","c9","s9","v5","c25","s25","v12","c6","s6","ex"]}
    for e in entries:
        b18=split_evenly(e["v18"],e["nbills"]); b5=split_evenly(e["v5"],e["nbills"])
        b12=split_evenly(e["v12"],e["nbills"]); bex=split_evenly(e["vex"],e["nbills"])
        for i in range(e["nbills"]):
            bv18=b18[i];bv5=b5[i];bv12=b12[i];bvex=bex[i]
            bval=r2(bv18+bv5+bv12+bvex)
            if bval==0: continue
            bgross=r2(bv18*1.18+bv5*1.05+bv12*1.12+bvex)
            ws.append([e["date"],"Cash","Cash","Sales",str(vno),None,None,
                bval,bgross,
                bv18 or None,r2(bv18*.09) or None,r2(bv18*.09) or None,
                bv5  or None,r2(bv5*.025) or None,r2(bv5*.025) or None,
                bv12 or None,r2(bv12*.06) or None,r2(bv12*.06) or None,
                bvex or None])
            g["val"]+=bval;g["gross"]+=bgross
            g["v18"]+=bv18;g["c9"]+=r2(bv18*.09);g["s9"]+=r2(bv18*.09)
            g["v5"]+=bv5;g["c25"]+=r2(bv5*.025);g["s25"]+=r2(bv5*.025)
            g["v12"]+=bv12;g["c6"]+=r2(bv12*.06);g["s6"]+=r2(bv12*.06)
            g["ex"]+=bvex; vno+=1
    ws.append([None,"Grand Total",None,None,None,None,None,
        r2(g["val"]),r2(g["gross"]),
        r2(g["v18"]) or None,r2(g["c9"]) or None,r2(g["s9"]) or None,
        r2(g["v5"])  or None,r2(g["c25"])or None,r2(g["s25"])or None,
        r2(g["v12"]) or None,r2(g["c6"]) or None,r2(g["s6"]) or None,
        r2(g["ex"])  or None])
    set_widths(ws,[12,14,12,12,10,14,16,10,10,10,13,13,10,14,14,10,13,13,13])
    return to_bytes(wb), g, vno-1

def make_purchase_xlsx(mk, purchases):
    m,y=parse_mk(mk)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Purchase Register"
    for row in header_block("Purchase Register",m,y): ws.append(row)
    ws.append(["Date","Particulars","Supplier","Voucher Type","Voucher No.",
        "Supplier Invoice No.","Supplier Invoice Date","GSTIN/UIN",
        "Value","Addl. Cost","Gross Total",
        "18 % Purchase","9% CGST","9% SGST","Round Off",
        "5 % Purchase","2.5% CGST","2.5 % SGST","EXAMPTED","CASH DISCOUNT"])
    g={k:0.0 for k in ["val","gross","v18","c9","s9","ro","v5","c25","s25","ex","disc"]}
    for i,p in enumerate(purchases,1):
        val=r2((p.get("val18") or 0)+(p.get("val5") or 0)+(p.get("val12") or 0)+(p.get("exempt") or 0))
        ws.append([p.get("inv_date_obj") or p.get("inv_date",""),
            p.get("supplier",""),p.get("supplier",""),"Purchase",str(p.get("pur_vno", i)),
            p.get("invno",""),p.get("inv_date_obj") or p.get("inv_date",""),p.get("gstin",""),
            val,None,r2(p.get("gross") or 0),
            p.get("val18") or None,p.get("cgst9") or None,p.get("sgst9") or None,p.get("round_off") or None,
            p.get("val5")  or None,p.get("cgst25")or None,p.get("sgst25")or None,
            p.get("exempt")or None,p.get("discount")or None])
        g["val"]+=val;g["gross"]+=p.get("gross") or 0
        g["v18"]+=p.get("val18") or 0;g["c9"]+=p.get("cgst9") or 0;g["s9"]+=p.get("sgst9") or 0
        g["v5"]+=p.get("val5") or 0;g["c25"]+=p.get("cgst25") or 0;g["s25"]+=p.get("sgst25") or 0
        g["ex"]+=p.get("exempt") or 0;g["ro"]+=p.get("round_off") or 0;g["disc"]+=p.get("discount") or 0
    ws.append([None,"Grand Total",None,None,None,None,None,None,
        r2(g["val"]),None,r2(g["gross"]),
        r2(g["v18"]) or None,r2(g["c9"]) or None,r2(g["s9"]) or None,r2(g["ro"]) or None,
        r2(g["v5"])  or None,r2(g["c25"])or None,r2(g["s25"])or None,
        r2(g["ex"])  or None,r2(g["disc"])or None])
    set_widths(ws,[12,20,20,10,10,18,16,18,12,8,12,14,10,10,10,12,10,10,10,12])
    return to_bytes(wb), g
