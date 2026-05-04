"""
================================================================================
Sri Lakshmi Venkateshwara Traders — Fertilizer Shop Management Tool
================================================================================
Tech Stack : Python 3 | Streamlit (UI) | SQLite (Local DB)
Libraries  : reportlab (PDF Invoice) | xlsxwriter (CA-Ready Excel Export)
GST Logic  : Intra-state — Variable GST per product (0%, 5%, 12%, 18%, 28%)
             Each rate is split equally into CGST + SGST (e.g. 5% → 2.5%+2.5%)
QR Code    : USB HID QR scanner support — scanner reads QR code and types the
             encoded text into the input field automatically.
             Lookup logic: if QR code exists → restock prompt
                           if QR code is new → new product registration form
Author     : Built for Sri Lakshmi Venkateshwara Traders
================================================================================
"""

import streamlit as st
import sqlite3
import io
import datetime
import xlsxwriter
import base64, json, re, random, calendar, requests
import openpyxl
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# ─────────────────────────────────────────────────────────────────────────────
# SHOP CONSTANTS  — edit these once
# ─────────────────────────────────────────────────────────────────────────────
DB_PATH         = "slv_traders.db"
SHOP_NAME       = "Sri Lakshmi Venkateshwara Traders"
SHOP_ADDRESS1   = "#20, Near Bank Of Baroda"
SHOP_ADDRESS2   = "Mysore Bangalore Expressway, Thubinakere"
SHOP_ADDRESS3   = "Mandya, Karnataka — 571402"
SHOP_PHONE      = "+91-9743007647"
SHOP_EMAIL      = "chandana4192@gmail.com"
SHOP_ADDR1      = SHOP_ADDRESS1 + ", " + SHOP_ADDRESS2
SHOP_ADDR2      = SHOP_ADDRESS3
SHOP_STATE      = "Karnataka — 29"
GSTIN           = "29CDTPB8883L1ZH"
STATE_CODE      = "29"                # 29 = Karnataka
LOW_STOCK_ALERT = 10                  # show alert below this qty

# ─────────────────────────────────────────────────────────────────────────────
# GST RATE MASTER
# Intra-state: total GST split 50:50 into CGST + SGST
#   0%  → 0%   + 0%    Exempt (seeds, organic manure)
#   5%  → 2.5% + 2.5%  Fertilizers Chapter 31
#  12%  → 6%   + 6%    Pesticides / insecticides Chapter 38
#  18%  → 9%   + 9%    Micronutrients, specialty chemicals
#  28%  → 14%  + 14%   Luxury / specialty chemicals
# ─────────────────────────────────────────────────────────────────────────────
GST_RATES = {
    "0%  — Exempt (Seeds/Organic)"   : 0.00,
    "5%  — Fertilizers (Ch.31)"      : 0.05,
    "12% — Pesticides (Ch.38)"       : 0.12,
    "18% — Micronutrients/Chemicals" : 0.18,
    "28% — Specialty Chemicals"      : 0.28,
}
DEFAULT_GST_LABEL = "5%  — Fertilizers (Ch.31)"

# ─────────────────────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────────────────────
def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Create tables, auto-migrate old DBs, seed sample data."""
    conn = get_conn()
    c    = conn.cursor()

    # ── Inventory ──────────────────────────────────────────────────────────
    c.execute("""
        CREATE TABLE IF NOT EXISTS inventory (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            qr_code     TEXT    UNIQUE,
            name        TEXT    NOT NULL,
            hsn_code    TEXT    NOT NULL,
            section     TEXT    NOT NULL,
            row_no      TEXT    NOT NULL,
            slot        TEXT    NOT NULL,
            unit        TEXT    DEFAULT 'Kg',
            quantity    REAL    DEFAULT 0,
            mrp         REAL    NOT NULL,
            cost_price  REAL    DEFAULT 0,
            gst_rate    REAL    DEFAULT 0.05
        )
    """)
    # Auto-migrate: add new columns to existing DBs without data loss
    # SQLite does NOT support ADD COLUMN with UNIQUE constraint directly.
    # Solution: add column without UNIQUE, then create a unique index separately.
    existing = [r[1] for r in c.execute("PRAGMA table_info(inventory)")]
    if "gst_rate" not in existing:
        c.execute("ALTER TABLE inventory ADD COLUMN gst_rate REAL DEFAULT 0.05")
    if "qr_code" not in existing:
        c.execute("ALTER TABLE inventory ADD COLUMN qr_code TEXT")
    # Unique index — safe to run every time
    c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_inv_qr ON inventory(qr_code) WHERE qr_code IS NOT NULL")

    # ── Invoices ───────────────────────────────────────────────────────────
    c.execute("""
        CREATE TABLE IF NOT EXISTS invoices (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_no      TEXT    UNIQUE NOT NULL,
            customer_name   TEXT,
            customer_phone  TEXT,
            customer_gstin  TEXT,
            invoice_date    TEXT    NOT NULL,
            taxable_value   REAL,
            cgst_amount     REAL,
            sgst_amount     REAL,
            total_amount    REAL
        )
    """)

    # ── Invoice Line Items ─────────────────────────────────────────────────
    c.execute("""
        CREATE TABLE IF NOT EXISTS invoice_items (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_no      TEXT    NOT NULL,
            product_id      INTEGER NOT NULL,
            product_name    TEXT,
            hsn_code        TEXT,
            quantity        REAL,
            unit            TEXT,
            unit_price      REAL,
            gst_rate        REAL,
            taxable_value   REAL,
            cgst_amount     REAL,
            sgst_amount     REAL,
            line_total      REAL,
            FOREIGN KEY (invoice_no) REFERENCES invoices(invoice_no)
        )
    """)
    existing_ii = [r[1] for r in c.execute("PRAGMA table_info(invoice_items)")]
    if "gst_rate" not in existing_ii:
        c.execute("ALTER TABLE invoice_items ADD COLUMN gst_rate REAL DEFAULT 0.05")

    # ── QR Code Registry ────────────────────────────────────────────────────
    # Tracks every scanned QR code and which product it maps to
    c.execute("""
        CREATE TABLE IF NOT EXISTS qr_registry (
            qr_code         TEXT    PRIMARY KEY,
            product_id      INTEGER,
            product_name    TEXT,
            first_scanned   TEXT,
            scan_count      INTEGER DEFAULT 1,
            FOREIGN KEY (product_id) REFERENCES inventory(id)
        )
    """)

    # ── Seed sample products (only if DB is fresh) ─────────────────────────
    c.execute("SELECT COUNT(*) FROM inventory")
    if c.fetchone()[0] == 0:
        # barcode, name, hsn, section, row, slot, unit, qty, mrp, cost, gst
        samples = [
            ("8901030870925","Urea (46% N)",          "31021010","Chemical Section","Row 1","AA","Bag",500, 380, 300,0.05),
            ("8901030870932","DAP (Diammonium Phos)", "31053000","Chemical Section","Row 1","AB","Bag",150,1350,1100,0.05),
            ("8901030870949","MOP (Muriate of Pot)",  "31042010","Chemical Section","Row 2","AA","Bag",200, 900, 750,0.05),
            ("8901030870956","NPK 10:26:26",          "31059010","Chemical Section","Row 2","AB","Bag",180,1200, 980,0.05),
            ("8901030870963","Gypsum",                "25201000","Mineral Section", "Row 3","BA","Bag",300, 250, 180,0.05),
            ("8901030870970","Zinc Sulphate",         "28330300","Micro-Nutrient",  "Row 3","BB","Kg",  85,  85,  65,0.18),
            ("8901030870987","Boron (Borax)",         "28402000","Micro-Nutrient",  "Row 4","CA","Kg", 120, 120,  95,0.18),
            ("8901030870994","Ferrous Sulphate",      "28332910","Micro-Nutrient",  "Row 4","CB","Kg",  60,  60,  45,0.18),
            ("8901030871007","Organic Compost",       "31010010","Organic Section", "Row 5","DA","Bag",350, 350, 250,0.00),
            ("8901030871014","Neem Cake",             "31010090","Organic Section", "Row 5","DB","Kg",  40,  40,  28,0.00),
            ("8901030871021","Potassium Nitrate",     "31043000","Chemical Section","Row 6","EA","Kg",  95,  95,  75,0.05),
            ("8901030871038","Calcium Nitrate",       "31023000","Chemical Section","Row 6","EB","Kg",  70,  70,  55,0.05),
            ("8901030871045","Chlorpyrifos 20% EC",   "38081091","Pesticide Section","Row 7","FA","Litre",45,450,350,0.12),
            ("8901030871052","Imidacloprid 17.8% SL", "38081099","Pesticide Section","Row 7","FB","Litre",60,600,480,0.12),
        ]
        c.executemany("""
            INSERT INTO inventory
                (qr_code,name,hsn_code,section,row_no,slot,unit,quantity,mrp,cost_price,gst_rate)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, samples)
        now = datetime.datetime.now().isoformat()
        for s in samples:
            c.execute("""
                INSERT OR IGNORE INTO qr_registry
                    (qr_code,product_id,product_name,first_scanned,scan_count)
                SELECT ?,id,name,?,1 FROM inventory WHERE qr_code=?
            """, (s[0], now, s[0]))

    conn.commit()
    conn.close()

# ─────────────────────────────────────────────────────────────────────────────
# QR CODE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def lookup_barcode(qr_code: str):
    """
    Look up a QR code in inventory.
    Returns (product_dict_or_None, is_new_product_bool)
    """
    qr_code = qr_code.strip()
    conn    = get_conn()
    prod    = conn.execute(
        "SELECT * FROM inventory WHERE qr_code=?", (qr_code,)
    ).fetchone()
    conn.close()
    if prod:
        return dict(prod), False
    return None, True


def register_barcode_scan(qr_code: str, product_id: int, product_name: str):
    """Log every QR scan to qr_registry for audit trail."""
    conn = get_conn()
    now  = datetime.datetime.now().isoformat()
    existing = conn.execute(
        "SELECT qr_code FROM qr_registry WHERE qr_code=?", (qr_code,)
    ).fetchone()
    if existing:
        conn.execute(
            "UPDATE qr_registry SET scan_count=scan_count+1 WHERE qr_code=?",
            (qr_code,)
        )
    else:
        conn.execute("""
            INSERT INTO qr_registry (qr_code,product_id,product_name,first_scanned,scan_count)
            VALUES (?,?,?,?,1)
        """, (qr_code, product_id, product_name, now))
    conn.commit()
    conn.close()

# ─────────────────────────────────────────────────────────────────────────────
# INVOICE NUMBER GENERATOR
# ─────────────────────────────────────────────────────────────────────────────
def generate_invoice_no():
    conn   = get_conn()
    now    = datetime.datetime.now()
    prefix = f"SLV-{now.strftime('%Y%m')}-"
    row    = conn.execute(
        "SELECT invoice_no FROM invoices WHERE invoice_no LIKE ? ORDER BY id DESC LIMIT 1",
        (prefix + "%",)
    ).fetchone()
    conn.close()
    seq = int(row["invoice_no"].split("-")[-1]) + 1 if row else 1
    return f"{prefix}{seq:04d}"

# ─────────────────────────────────────────────────────────────────────────────
# GST CALCULATION — variable rate per product
# ─────────────────────────────────────────────────────────────────────────────
def calculate_gst(taxable_value: float, gst_rate: float = 0.05):
    """
    Intra-state GST: total rate split 50:50 → CGST + SGST
    e.g. 12% total → CGST 6% + SGST 6%
    """
    half  = gst_rate / 2
    cgst  = round(taxable_value * half, 2)
    sgst  = round(taxable_value * half, 2)
    total = round(taxable_value + cgst + sgst, 2)
    return cgst, sgst, total

# ─────────────────────────────────────────────────────────────────────────────
# PDF INVOICE GENERATOR  — professional, no overlaps
# ─────────────────────────────────────────────────────────────────────────────
def generate_pdf_invoice(invoice_data: dict, items: list) -> bytes:
    buffer     = io.BytesIO()
    doc        = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=8*mm,   bottomMargin=12*mm
    )

    # ── Colours ─────────────────────────────────────────────────────────────
    brand_blue  = colors.HexColor("#1a3c5e")
    brand_green = colors.HexColor("#2e7d32")
    light_grey  = colors.HexColor("#f5f5f5")
    mid_grey    = colors.HexColor("#e0e0e0")
    white       = colors.white

    # ── Styles ──────────────────────────────────────────────────────────────
    s_shop  = ParagraphStyle("shop",  fontSize=16, textColor=white,
                              alignment=TA_CENTER, fontName="Helvetica-Bold",
                              leading=20)
    s_addr  = ParagraphStyle("addr",  fontSize=8,  textColor=colors.HexColor("#c8e6c9"),
                              alignment=TA_CENTER, fontName="Helvetica", leading=12)
    s_label = ParagraphStyle("lbl",   fontSize=8,  textColor=brand_blue,
                              fontName="Helvetica-Bold")
    s_val   = ParagraphStyle("val",   fontSize=8,  fontName="Helvetica")
    s_foot  = ParagraphStyle("ft",    fontSize=7,  textColor=colors.grey,
                              alignment=TA_CENTER, fontName="Helvetica-Oblique")
    s_ti    = ParagraphStyle("ti",    fontSize=9,  textColor=white,
                              alignment=TA_CENTER, fontName="Helvetica-Bold", leading=14)

    story = []
    inv   = invoice_data

    # ══════════════════════════════════════════════════════════════════════
    # HEADER BAND — shop name + address on dark blue background
    # Uses a single-cell Table so background fills edge to edge
    # ══════════════════════════════════════════════════════════════════════
    header_content = [
        [Paragraph(SHOP_NAME, s_shop)],
        [Paragraph(SHOP_ADDRESS1, s_addr)],
        [Paragraph(SHOP_ADDRESS2, s_addr)],
        [Paragraph(SHOP_ADDRESS3, s_addr)],
        [Paragraph(
            f"📞 {SHOP_PHONE}   |   ✉ {SHOP_EMAIL}   |   GSTIN: {GSTIN}",
            s_addr
        )],
    ]
    t_banner = Table(header_content, colWidths=[186*mm])
    t_banner.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), brand_blue),
        ("TOPPADDING",    (0,0),(-1,-1), 6),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("LEFTPADDING",   (0,0),(-1,-1), 8),
        ("RIGHTPADDING",  (0,0),(-1,-1), 8),
        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
    ]))
    story.append(t_banner)
    story.append(Spacer(1, 2*mm))

    # TAX INVOICE label + Invoice No + Date on one row
    t_title = Table([[
        Paragraph("TAX INVOICE", s_ti),
        Paragraph(f"Invoice No: <b>{inv['invoice_no']}</b>", s_val),
        Paragraph(f"Date: <b>{inv['invoice_date'][:10]}</b>", s_val),
    ]], colWidths=[80*mm, 60*mm, 46*mm])
    t_title.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(0,0), brand_green),
        ("BACKGROUND",    (1,0),(2,0), light_grey),
        ("ALIGN",         (0,0),(0,0), "CENTER"),
        ("ALIGN",         (1,0),(2,0), "LEFT"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("FONTSIZE",      (0,0),(-1,-1), 8),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LEFTPADDING",   (0,0),(-1,-1), 6),
        ("BOX",           (0,0),(-1,-1), 0.5, mid_grey),
        ("LINEAFTER",     (0,0),(1,0),   0.5, mid_grey),
    ]))
    story.append(t_title)
    story.append(Spacer(1, 3*mm))

    # ══════════════════════════════════════════════════════════════════════
    # SELLER  |  BUYER  — side by side, no overlap
    # ══════════════════════════════════════════════════════════════════════
    def cell(label, value):
        return Paragraph(f"<b>{label}:</b> {value}", s_val)

    seller_rows = [
        [Paragraph("<b>Sold By</b>", s_label)],
        [cell("Name",  SHOP_NAME)],
        [cell("Addr",  f"{SHOP_ADDRESS1}, {SHOP_ADDRESS2}")],
        [cell("",      SHOP_ADDRESS3)],
        [cell("GSTIN", GSTIN)],
        [cell("State", f"Karnataka ({STATE_CODE})")],
    ]
    buyer_name  = inv.get("customer_name")  or "Cash Customer"
    buyer_phone = inv.get("customer_phone") or "—"
    buyer_gstin = inv.get("customer_gstin") or "Unregistered"
    buyer_rows  = [
        [Paragraph("<b>Bill To</b>", s_label)],
        [cell("Name",  buyer_name)],
        [cell("Phone", buyer_phone)],
        [cell("GSTIN", buyer_gstin)],
    ]

    t_seller = Table(seller_rows, colWidths=[88*mm])
    t_seller.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(0,0), light_grey),
        ("BOX",           (0,0),(-1,-1), 0.5, mid_grey),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 2),
        ("LEFTPADDING",   (0,0),(-1,-1), 6),
    ]))

    t_buyer = Table(buyer_rows, colWidths=[88*mm])
    t_buyer.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(0,0), light_grey),
        ("BOX",           (0,0),(-1,-1), 0.5, mid_grey),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 2),
        ("LEFTPADDING",   (0,0),(-1,-1), 6),
    ]))

    t_parties = Table([[t_seller, Spacer(10*mm,1), t_buyer]],
                      colWidths=[88*mm, 10*mm, 88*mm])
    t_parties.setStyle(TableStyle([
        ("VALIGN", (0,0),(-1,-1), "TOP"),
    ]))
    story.append(t_parties)
    story.append(Spacer(1, 4*mm))

    # ══════════════════════════════════════════════════════════════════════
    # LINE ITEMS TABLE
    # ══════════════════════════════════════════════════════════════════════
    col_w   = [6*mm, 46*mm, 14*mm, 9*mm, 9*mm, 9*mm, 16*mm, 18*mm, 16*mm, 16*mm, 17*mm]
    headers = ["#", "Product", "HSN", "GST%", "Qty", "Unit",
               "Rate(₹)", "Taxable(₹)", "CGST(₹)", "SGST(₹)", "Total(₹)"]
    table_data = [headers]
    for i, item in enumerate(items, 1):
        gp   = (item.get("gst_rate", 0.05) or 0) * 100
        half = gp / 2
        table_data.append([
            str(i),
            Paragraph(item["product_name"],
                      ParagraphStyle("pn", fontSize=7, fontName="Helvetica",
                                     leading=9, wordWrap="LTR")),
            item["hsn_code"],
            f"{gp:.0f}%",
            str(int(item["quantity"]) if item["quantity"] == int(item["quantity"])
                else item["quantity"]),
            item["unit"],
            f"{item['unit_price']:.2f}",
            f"{item['taxable_value']:.2f}",
            f"{item['cgst_amount']:.2f}",
            f"{item['sgst_amount']:.2f}",
            f"{item['line_total']:.2f}",
        ])

    t_items = Table(table_data, colWidths=col_w, repeatRows=1)
    t_items.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  brand_blue),
        ("TEXTCOLOR",     (0,0),(-1,0),  white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 7),
        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
        ("ALIGN",         (1,1),(1,-1),  "LEFT"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [white, light_grey]),
        ("GRID",          (0,0),(-1,-1), 0.25, mid_grey),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LEFTPADDING",   (0,0),(-1,-1), 3),
        ("RIGHTPADDING",  (0,0),(-1,-1), 3),
    ]))
    story.append(t_items)
    story.append(Spacer(1, 3*mm))

    # ══════════════════════════════════════════════════════════════════════
    # TOTALS — right-aligned block
    # ══════════════════════════════════════════════════════════════════════
    tv   = inv["taxable_value"]
    cgst = inv["cgst_amount"]
    sgst = inv["sgst_amount"]
    tot  = inv["total_amount"]

    s_tot_lbl = ParagraphStyle("tl", fontSize=8,  fontName="Helvetica",
                                alignment=TA_RIGHT)
    s_tot_val = ParagraphStyle("tv", fontSize=8,  fontName="Helvetica",
                                alignment=TA_RIGHT)
    s_grand   = ParagraphStyle("gv", fontSize=11, fontName="Helvetica-Bold",
                                alignment=TA_RIGHT, textColor=brand_blue)

    # Build totals rows — include discount row only when discount > 0
    discount_amt = inv.get("discount_amt", 0) or 0
    sub_total    = inv.get("sub_total", tot)   or tot

    totals_rows = [
        [Paragraph("Taxable Value", s_tot_lbl), Paragraph(f"₹ {tv:.2f}",   s_tot_val)],
        [Paragraph("CGST",          s_tot_lbl), Paragraph(f"₹ {cgst:.2f}", s_tot_val)],
        [Paragraph("SGST",          s_tot_lbl), Paragraph(f"₹ {sgst:.2f}", s_tot_val)],
    ]
    if discount_amt > 0:
        totals_rows.append([
            Paragraph("Subtotal",   s_tot_lbl), Paragraph(f"₹ {sub_total:.2f}", s_tot_val)
        ])
        s_disc = ParagraphStyle("disc", fontSize=8, fontName="Helvetica-Bold",
                                 alignment=TA_RIGHT,
                                 textColor=colors.HexColor("#c62828"))
        totals_rows.append([
            Paragraph("Discount",   s_disc), Paragraph(f"- ₹ {discount_amt:.2f}", s_disc)
        ])
    totals_rows.append([
        Paragraph("<b>GRAND TOTAL</b>", s_grand),
        Paragraph(f"<b>₹ {tot:.2f}</b>", s_grand)
    ])

    t_totals = Table(totals_rows, colWidths=[120*mm, 66*mm])
    disc_row_idx = len(totals_rows) - 1  # last row is grand total
    style_cmds = [
        ("ALIGN",         (0,0),(-1,-1), "RIGHT"),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LINEABOVE",     (0,-1),(-1,-1), 1, brand_blue),
        ("LINEBELOW",     (0,-1),(-1,-1), 1, brand_blue),
        ("BACKGROUND",    (0,-1),(-1,-1), light_grey),
    ]
    t_totals.setStyle(TableStyle(style_cmds))
    story.append(t_totals)
    story.append(Spacer(1, 5*mm))

    # ══════════════════════════════════════════════════════════════════════
    # FOOTER
    # ══════════════════════════════════════════════════════════════════════
    t_footer = Table([[
        Paragraph("This is a computer-generated invoice. No signature required.", s_foot),
        Paragraph("<b>Authorised Signatory</b><br/>" + SHOP_NAME,
                  ParagraphStyle("sig", fontSize=8, fontName="Helvetica",
                                 alignment=TA_RIGHT, leading=12)),
    ]], colWidths=[120*mm, 66*mm])
    t_footer.setStyle(TableStyle([
        ("VALIGN",     (0,0),(-1,-1), "BOTTOM"),
        ("LINEABOVE",  (1,0),(1,0),   0.5, mid_grey),
        ("TOPPADDING", (0,0),(-1,-1), 4),
    ]))
    story.append(t_footer)

    doc.build(story)
    return buffer.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL GSTR-1 REPORT
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel_report() -> bytes:
    conn   = get_conn()
    today  = datetime.date.today()
    buffer = io.BytesIO()
    wb     = xlsxwriter.Workbook(buffer, {"in_memory": True})

    hdr_fmt   = wb.add_format({"bold":True,"bg_color":"#1a3c5e","font_color":"white",
                                "border":1,"align":"center","valign":"vcenter","text_wrap":True})
    cell_fmt  = wb.add_format({"border":1,"align":"left"})
    num_fmt   = wb.add_format({"border":1,"num_format":"#,##0.00","align":"right"})
    title_fmt = wb.add_format({"bold":True,"font_size":14,"font_color":"#1a3c5e"})
    sub_fmt   = wb.add_format({"italic":True,"font_color":"#555555"})
    total_fmt = wb.add_format({"bold":True,"bg_color":"#e8f5e9","border":1,
                                "num_format":"#,##0.00","align":"right"})

    COLS = [
        ("Date",15),("Invoice No",18),("Customer",25),("GSTIN",20),
        ("HSN",12),("Product",30),("GST Rate",10),("Qty",8),("Unit",8),
        ("Taxable Value",15),("CGST",13),("SGST",13),("Total",15),
    ]

    def write_sheet(name, where_sql, params):
        ws = wb.add_worksheet(name)
        ws.merge_range("A1:M1", f"{SHOP_NAME} — {name} Sales Report", title_fmt)
        ws.merge_range("A2:M2",
            f"GSTIN: {GSTIN}  |  Generated: {today:%d %b %Y}  |  For GSTR-1 Filing",
            sub_fmt)
        ws.set_row(0, 22)
        ws.set_row(2, 30)
        for ci, (cn, cw) in enumerate(COLS):
            ws.write(2, ci, cn, hdr_fmt)
            ws.set_column(ci, ci, cw)

        rows = conn.execute(f"""
            SELECT i.invoice_date, i.invoice_no, i.customer_name, i.customer_gstin,
                   ii.hsn_code, ii.product_name, ii.gst_rate,
                   ii.quantity, ii.unit,
                   ii.taxable_value, ii.cgst_amount, ii.sgst_amount, ii.line_total
            FROM invoices i JOIN invoice_items ii ON i.invoice_no=ii.invoice_no
            WHERE {where_sql}
            ORDER BY i.invoice_date, i.invoice_no
        """, params).fetchall()

        ri     = 3
        # Use named columns for totals — avoids index offset bugs
        totals = {"taxable_value": 0.0, "cgst_amount": 0.0,
                  "sgst_amount": 0.0,   "line_total":  0.0}
        for r in rows:
            ws.write(ri, 0,  r["invoice_date"],                          cell_fmt)
            ws.write(ri, 1,  r["invoice_no"],                            cell_fmt)
            ws.write(ri, 2,  r["customer_name"]  or "Cash Customer",     cell_fmt)
            ws.write(ri, 3,  r["customer_gstin"] or "Unregistered",      cell_fmt)
            ws.write(ri, 4,  r["hsn_code"],                              cell_fmt)
            ws.write(ri, 5,  r["product_name"],                          cell_fmt)
            ws.write(ri, 6,  f"{(r['gst_rate'] or 0)*100:.0f}%",        cell_fmt)
            ws.write(ri, 7,  r["quantity"],                              num_fmt)
            ws.write(ri, 8,  r["unit"],                                  cell_fmt)
            ws.write(ri, 9,  r["taxable_value"],                         num_fmt)
            ws.write(ri, 10, r["cgst_amount"],                           num_fmt)
            ws.write(ri, 11, r["sgst_amount"],                           num_fmt)
            ws.write(ri, 12, r["line_total"],                            num_fmt)
            # Accumulate totals using column names — safe, no index arithmetic
            totals["taxable_value"] += r["taxable_value"] or 0
            totals["cgst_amount"]   += r["cgst_amount"]   or 0
            totals["sgst_amount"]   += r["sgst_amount"]   or 0
            totals["line_total"]    += r["line_total"]    or 0
            ri += 1

        if rows:
            ws.merge_range(ri, 0, ri, 8, "TOTAL", hdr_fmt)
            ws.write(ri, 9,  totals["taxable_value"], total_fmt)
            ws.write(ri, 10, totals["cgst_amount"],   total_fmt)
            ws.write(ri, 11, totals["sgst_amount"],   total_fmt)
            ws.write(ri, 12, totals["line_total"],    total_fmt)
        else:
            ws.merge_range(ri, 0, ri, 12, "No transactions for this period.", cell_fmt)
        ws.freeze_panes(3, 0)

    today_s    = today.isoformat()
    week_start = today - datetime.timedelta(days=today.weekday())
    week_end   = week_start + datetime.timedelta(days=6)

    write_sheet("Daily",   "DATE(i.invoice_date)=?",               (today_s,))
    write_sheet("Weekly",  "DATE(i.invoice_date) BETWEEN ? AND ?", (week_start.isoformat(), week_end.isoformat()))
    write_sheet("Monthly", "strftime('%Y-%m',i.invoice_date)=?",   (today.strftime("%Y-%m"),))

    conn.close()
    wb.close()
    return buffer.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT CONFIG & CSS
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SLV Traders — Shop Manager",
    page_icon="🌱", layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
.main-header {
    background: linear-gradient(135deg,#1a3c5e,#2e7d32);
    color:white; padding:1.2rem 1.5rem; border-radius:10px; margin-bottom:1rem;
}
.main-header h1 { color:white; margin:0; font-size:1.7rem; }
.main-header p  { color:#c8e6c9; margin:0; font-size:0.85rem; }
.scan-found {
    background:#e8f5e9; border-left:6px solid #2e7d32;
    padding:14px 18px; border-radius:8px; margin:10px 0;
}
.scan-new {
    background:#fff8e1; border-left:6px solid #f9a825;
    padding:14px 18px; border-radius:8px; margin:10px 0;
}
.scan-new h3   { color:#e65100; margin:0 0 6px 0; }
.scan-found h3 { color:#1b5e20; margin:0 0 6px 0; }
.detail-row { display:flex; gap:12px; flex-wrap:wrap; margin-top:8px; }
.detail-pill {
    background:white; border:1px solid #ccc;
    padding:4px 12px; border-radius:20px; font-size:0.85rem;
}

/* ── Sidebar background ── */
div[data-testid="stSidebar"] { background:#1a3c5e !important; }
div[data-testid="stSidebar"] * { color:#e3f2fd !important; }

/* ── Nav buttons: full-width, hover highlight ── */
div[data-testid="stSidebar"] .stButton > button {
    width: 100%;
    text-align: left !important;
    background: transparent !important;
    color: #c8e6c9 !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 10px 16px !important;
    font-size: 0.97rem !important;
    font-weight: 400 !important;
    margin-bottom: 2px !important;
    transition: background 0.15s ease, color 0.15s ease !important;
    cursor: pointer !important;
}
div[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(255,255,255,0.12) !important;
    color: #ffffff !important;
}
/* Active page button — bright white + left accent bar */
div[data-testid="stSidebar"] .stButton > button[data-active="true"],
div[data-testid="stSidebar"] .nav-active > button {
    background: rgba(255,255,255,0.18) !important;
    color: #ffffff !important;
    font-weight: 600 !important;
    border-left: 4px solid #66bb6a !important;
    padding-left: 12px !important;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE INITIALISATION
# ─────────────────────────────────────────────────────────────────────────────
for key, val in {
    "cart":             [],
    "last_qr_code":     "",
    "qr_result":        None,
    "qr_is_new":        False,
}.items():
    if key not in st.session_state:
        st.session_state[key] = val

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR NAVIGATION
# ─────────────────────────────────────────────────────────────────────────────
# ── Initialise page in session state ─────────────────────────────────────
if "page" not in st.session_state:
    st.session_state.page = "🏠 Dashboard"

NAV_PAGES = [
    ("🏠", "Dashboard"),
    ("📦", "Inventory"),
    ("🔬", "QR Scanner"),
    ("🧾", "New Bill"),
    ("📊", "Reports"),
    ("🔍", "Search Product"),
    ("📋", "GST Registers"),
]

with st.sidebar:
    st.markdown("""
    <div style="padding:18px 16px 8px 16px;">
        <div style="font-size:1.5rem;font-weight:700;color:white;letter-spacing:0.5px;">
            🌱 SLV Traders
        </div>
        <div style="font-size:0.75rem;color:#a5d6a7;margin-top:2px;">
            Shop Management System
        </div>
    </div>
    <hr style="border-color:rgba(255,255,255,0.15);margin:8px 0 12px 0;">
    """, unsafe_allow_html=True)

    for icon, label in NAV_PAGES:
        full = f"{icon} {label}"
        is_active = st.session_state.page == full
        # Wrap active button in a div with class for CSS targeting
        if is_active:
            st.markdown('<div class="nav-active">', unsafe_allow_html=True)
        if st.button(full, key=f"nav_{label}", use_container_width=True):
            st.session_state.page = full
            st.rerun()
        if is_active:
            st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("""
    <hr style="border-color:rgba(255,255,255,0.15);margin:12px 0 10px 0;">
    """, unsafe_allow_html=True)
    st.markdown(f"<div style='font-size:0.78rem;color:#a5d6a7;padding:0 16px;'>"
                f"<b>GSTIN:</b> {GSTIN}<br>"
                f"<b>Date:</b> {datetime.date.today().strftime('%d %b %Y')}"
                f"</div>", unsafe_allow_html=True)

page = st.session_state.page

# ─────────────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="main-header">
    <h1>🌱 Sri Lakshmi Venkateshwara Traders</h1>
    <p>Fertilizer Shop Management System &nbsp;|&nbsp; GSTIN: {GSTIN}</p>
</div>
""", unsafe_allow_html=True)

init_db()
conn = get_conn()

# ═════════════════════════════════════════════════════════════════════════════
# PAGE — DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════
# GST REGISTERS MODULE
# ═══════════════════════════════════════════════════════════════

MON_ABBR = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
             7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

MONTHS = [
    ("11-2025", "November 2025"),
    ("12-2025", "December 2025"),
    ("01-2026", "January 2026"),
    ("02-2026", "February 2026"),
    ("03-2026", "March 2026"),
    ("04-2026", "April 2026"),
]
MONTH_FP = {
    "11-2025":"112025","12-2025":"122025",
    "01-2026":"012026","02-2026":"022026","03-2026":"032026","04-2026":"042026",
}
OCT_LAST_VNO = 474          # October sales ended at voucher 474
OCT_LAST_PUR_VNO = 61   # October purchase ended at voucher 61

# ── Voucher start numbers — sourced directly from filed GSTR-1 JSONs ─────────
# Oct: from=351, to=474 → Nov starts at 475
# Nov: from=475, to=591 → Dec starts at 592
# Dec: from=592, to=710 → Jan starts at 711 (filed 714 — 3 cancelled before filing)
# Jan: from=714, to=830 → Feb starts at 831 (filed 834 — 3 cancelled before filing)
# Feb: from=834, to=939 → Mar starts at 940
VOUCHER_START = {
    "11-2025": 475,
    "12-2025": 592,
    "01-2026": 714,
    "02-2026": 834,
    "03-2026": 940,
    "04-2026": 1071,
}

# ── Pre-loaded HSN data from filed GSTR-1 JSONs (Nov 2025 – Feb 2026) ────────
# Each entry: (num, hsn_sc, desc, uqc, qty, rt, txval, camt, samt, csamt)
# These are the exact values accepted by the GST portal — do not modify.
FILED_HSN = {
    "11-2025": [
        {"num":1,"hsn_sc":"1006","desc":"","user_desc":"","uqc":"PAC","qty":9,"rt":0,"txval":65740.0,"iamt":0,"camt":0.0,"samt":0.0,"csamt":0},
        {"num":2,"hsn_sc":"31059010","desc":"","user_desc":"","uqc":"PCS","qty":1,"rt":5,"txval":1850.12,"iamt":0,"camt":46.25,"samt":46.25,"csamt":0},
        {"num":3,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PCS","qty":1,"rt":5,"txval":1526.47,"iamt":0,"camt":38.16,"samt":38.16,"csamt":0},
        {"num":4,"hsn_sc":"38089349","desc":"","user_desc":"","uqc":"PAC","qty":1,"rt":5,"txval":1454.01,"iamt":0,"camt":36.35,"samt":36.35,"csamt":0},
        {"num":5,"hsn_sc":"28331990","desc":"","user_desc":"","uqc":"PCS","qty":1,"rt":12,"txval":1374.5,"iamt":0,"camt":82.47,"samt":82.47,"csamt":0},
        {"num":6,"hsn_sc":"38089290","desc":"","user_desc":"","uqc":"PCS","qty":686,"rt":18,"txval":168968.8,"iamt":0,"camt":15207.19,"samt":15207.19,"csamt":0},
        {"num":7,"hsn_sc":"38089390","desc":"","user_desc":"","uqc":"PCS","qty":165,"rt":18,"txval":27430.0,"iamt":0,"camt":2468.7,"samt":2468.7,"csamt":0},
        {"num":8,"hsn_sc":"38089390","desc":"","user_desc":"","uqc":"PAC","qty":52,"rt":18,"txval":14537.9,"iamt":0,"camt":1308.41,"samt":1308.41,"csamt":0},
        {"num":9,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PCS","qty":96,"rt":18,"txval":23864.1,"iamt":0,"camt":2147.77,"samt":2147.77,"csamt":0},
        {"num":10,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PAC","qty":5,"rt":18,"txval":7954.7,"iamt":0,"camt":715.92,"samt":715.92,"csamt":0},
        {"num":11,"hsn_sc":"39201012","desc":"","user_desc":"","uqc":"PCS","qty":19,"rt":18,"txval":12617.8,"iamt":0,"camt":1135.6,"samt":1135.6,"csamt":0},
        {"num":12,"hsn_sc":"38089350","desc":"","user_desc":"","uqc":"PCS","qty":27,"rt":18,"txval":6583.2,"iamt":0,"camt":592.49,"samt":592.49,"csamt":0},
        {"num":13,"hsn_sc":"34029099","desc":"","user_desc":"","uqc":"PCS","qty":58,"rt":18,"txval":5760.3,"iamt":0,"camt":518.43,"samt":518.43,"csamt":0},
        {"num":14,"hsn_sc":"3808","desc":"","user_desc":"","uqc":"PCS","qty":19,"rt":18,"txval":3840.2,"iamt":0,"camt":345.62,"samt":345.62,"csamt":0},
        {"num":15,"hsn_sc":"38089910","desc":"","user_desc":"","uqc":"PCS","qty":3,"rt":18,"txval":1920.1,"iamt":0,"camt":172.81,"samt":172.81,"csamt":0},
        {"num":16,"hsn_sc":"380892","desc":"","user_desc":"","uqc":"PCS","qty":14,"rt":18,"txval":1097.2,"iamt":0,"camt":98.75,"samt":98.75,"csamt":0},
    ],
    "12-2025": [
        {"num":1,"hsn_sc":"31059010","desc":"","user_desc":"","uqc":"PCS","qty":12,"rt":5,"txval":21984.58,"iamt":0,"camt":549.61,"samt":549.61,"csamt":0},
        {"num":2,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PCS","qty":3,"rt":5,"txval":18138.71,"iamt":0,"camt":453.47,"samt":453.47,"csamt":0},
        {"num":3,"hsn_sc":"38089349","desc":"","user_desc":"","uqc":"PAC","qty":12,"rt":5,"txval":17277.69,"iamt":0,"camt":431.94,"samt":431.94,"csamt":0},
        {"num":4,"hsn_sc":"28331990","desc":"","user_desc":"","uqc":"PCS","qty":1,"rt":12,"txval":281.02,"iamt":0,"camt":16.86,"samt":16.86,"csamt":0},
        {"num":5,"hsn_sc":"38089290","desc":"","user_desc":"","uqc":"PCS","qty":282,"rt":18,"txval":69392.4,"iamt":0,"camt":6245.32,"samt":6245.32,"csamt":0},
        {"num":6,"hsn_sc":"38089390","desc":"","user_desc":"","uqc":"PCS","qty":68,"rt":18,"txval":11265.0,"iamt":0,"camt":1013.85,"samt":1013.85,"csamt":0},
        {"num":7,"hsn_sc":"38089390","desc":"","user_desc":"","uqc":"PAC","qty":21,"rt":18,"txval":5970.45,"iamt":0,"camt":537.34,"samt":537.34,"csamt":0},
        {"num":8,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PCS","qty":39,"rt":18,"txval":9800.55,"iamt":0,"camt":882.05,"samt":882.05,"csamt":0},
        {"num":9,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PAC","qty":2,"rt":18,"txval":3266.85,"iamt":0,"camt":294.02,"samt":294.02,"csamt":0},
        {"num":10,"hsn_sc":"39201012","desc":"","user_desc":"","uqc":"PCS","qty":8,"rt":18,"txval":5181.9,"iamt":0,"camt":466.37,"samt":466.37,"csamt":0},
        {"num":11,"hsn_sc":"38089350","desc":"","user_desc":"","uqc":"PCS","qty":11,"rt":18,"txval":2703.6,"iamt":0,"camt":243.32,"samt":243.32,"csamt":0},
        {"num":12,"hsn_sc":"34029099","desc":"","user_desc":"","uqc":"PCS","qty":24,"rt":18,"txval":2365.65,"iamt":0,"camt":212.91,"samt":212.91,"csamt":0},
        {"num":13,"hsn_sc":"3808","desc":"","user_desc":"","uqc":"PCS","qty":8,"rt":18,"txval":1577.1,"iamt":0,"camt":141.94,"samt":141.94,"csamt":0},
        {"num":14,"hsn_sc":"38089910","desc":"","user_desc":"","uqc":"PCS","qty":1,"rt":18,"txval":788.55,"iamt":0,"camt":70.97,"samt":70.97,"csamt":0},
        {"num":15,"hsn_sc":"380892","desc":"","user_desc":"","uqc":"PCS","qty":6,"rt":18,"txval":450.6,"iamt":0,"camt":40.55,"samt":40.55,"csamt":0},
    ],
    "01-2026": [
        {"num":1,"hsn_sc":"31059010","desc":"","user_desc":"","uqc":"PCS","qty":12,"rt":5,"txval":21610.1,"iamt":0,"camt":540.25,"samt":540.25,"csamt":0},
        {"num":2,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PCS","qty":3,"rt":5,"txval":17829.75,"iamt":0,"camt":445.74,"samt":445.74,"csamt":0},
        {"num":3,"hsn_sc":"38089349","desc":"","user_desc":"","uqc":"PAC","qty":12,"rt":5,"txval":16983.4,"iamt":0,"camt":424.58,"samt":424.58,"csamt":0},
        {"num":4,"hsn_sc":"28331990","desc":"","user_desc":"","uqc":"PCS","qty":1,"rt":12,"txval":798.7,"iamt":0,"camt":47.92,"samt":47.92,"csamt":0},
        {"num":5,"hsn_sc":"38089290","desc":"","user_desc":"","uqc":"PCS","qty":997,"rt":18,"txval":245599.2,"iamt":0,"camt":22103.93,"samt":22103.93,"csamt":0},
        {"num":6,"hsn_sc":"38089390","desc":"","user_desc":"","uqc":"PCS","qty":239,"rt":18,"txval":39870.0,"iamt":0,"camt":3588.3,"samt":3588.3,"csamt":0},
        {"num":7,"hsn_sc":"38089390","desc":"","user_desc":"","uqc":"PAC","qty":76,"rt":18,"txval":21131.1,"iamt":0,"camt":1901.8,"samt":1901.8,"csamt":0},
        {"num":8,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PCS","qty":140,"rt":18,"txval":34686.9,"iamt":0,"camt":3121.82,"samt":3121.82,"csamt":0},
        {"num":9,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PAC","qty":8,"rt":18,"txval":11562.3,"iamt":0,"camt":1040.61,"samt":1040.61,"csamt":0},
        {"num":10,"hsn_sc":"39201012","desc":"","user_desc":"","uqc":"PCS","qty":28,"rt":18,"txval":18340.2,"iamt":0,"camt":1650.62,"samt":1650.62,"csamt":0},
        {"num":11,"hsn_sc":"38089350","desc":"","user_desc":"","uqc":"PCS","qty":40,"rt":18,"txval":9568.8,"iamt":0,"camt":861.19,"samt":861.19,"csamt":0},
        {"num":12,"hsn_sc":"34029099","desc":"","user_desc":"","uqc":"PCS","qty":84,"rt":18,"txval":8372.7,"iamt":0,"camt":753.54,"samt":753.54,"csamt":0},
        {"num":13,"hsn_sc":"3808","desc":"","user_desc":"","uqc":"PCS","qty":28,"rt":18,"txval":5581.8,"iamt":0,"camt":502.36,"samt":502.36,"csamt":0},
        {"num":14,"hsn_sc":"38089910","desc":"","user_desc":"","uqc":"PCS","qty":4,"rt":18,"txval":2790.9,"iamt":0,"camt":251.18,"samt":251.18,"csamt":0},
        {"num":15,"hsn_sc":"380892","desc":"","user_desc":"","uqc":"PCS","qty":20,"rt":18,"txval":1594.8,"iamt":0,"camt":143.53,"samt":143.53,"csamt":0},
    ],
    "02-2026": [
        {"num":1,"hsn_sc":"1006","desc":"","user_desc":"","uqc":"PAC","qty":1,"rt":0,"txval":6920.0,"iamt":0,"camt":0.0,"samt":0.0,"csamt":0},
        {"num":2,"hsn_sc":"31059010","desc":"","user_desc":"","uqc":"PCS","qty":30,"rt":5,"txval":55366.38,"iamt":0,"camt":1384.16,"samt":1384.16,"csamt":0},
        {"num":3,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PCS","qty":7,"rt":5,"txval":45680.88,"iamt":0,"camt":1142.02,"samt":1142.02,"csamt":0},
        {"num":4,"hsn_sc":"38089349","desc":"","user_desc":"","uqc":"PAC","qty":30,"rt":5,"txval":43512.48,"iamt":0,"camt":1087.81,"samt":1087.81,"csamt":0},
        {"num":5,"hsn_sc":"28331990","desc":"","user_desc":"","uqc":"PCS","qty":1,"rt":12,"txval":123.72,"iamt":0,"camt":7.42,"samt":7.42,"csamt":0},
        {"num":6,"hsn_sc":"38089290","desc":"","user_desc":"","uqc":"PCS","qty":87,"rt":18,"txval":21344.4,"iamt":0,"camt":1921.0,"samt":1921.0,"csamt":0},
        {"num":7,"hsn_sc":"38089390","desc":"","user_desc":"","uqc":"PCS","qty":21,"rt":18,"txval":3465.0,"iamt":0,"camt":311.85,"samt":311.85,"csamt":0},
        {"num":8,"hsn_sc":"38089390","desc":"","user_desc":"","uqc":"PAC","qty":7,"rt":18,"txval":1836.45,"iamt":0,"camt":165.28,"samt":165.28,"csamt":0},
        {"num":9,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PCS","qty":12,"rt":18,"txval":3014.55,"iamt":0,"camt":271.31,"samt":271.31,"csamt":0},
        {"num":10,"hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PAC","qty":1,"rt":18,"txval":1004.85,"iamt":0,"camt":90.44,"samt":90.44,"csamt":0},
        {"num":11,"hsn_sc":"39201012","desc":"","user_desc":"","uqc":"PCS","qty":2,"rt":18,"txval":1593.9,"iamt":0,"camt":143.45,"samt":143.45,"csamt":0},
        {"num":12,"hsn_sc":"38089350","desc":"","user_desc":"","uqc":"PCS","qty":3,"rt":18,"txval":831.6,"iamt":0,"camt":74.84,"samt":74.84,"csamt":0},
        {"num":13,"hsn_sc":"34029099","desc":"","user_desc":"","uqc":"PCS","qty":7,"rt":18,"txval":727.65,"iamt":0,"camt":65.49,"samt":65.49,"csamt":0},
        {"num":14,"hsn_sc":"3808","desc":"","user_desc":"","uqc":"PCS","qty":2,"rt":18,"txval":485.1,"iamt":0,"camt":43.66,"samt":43.66,"csamt":0},
        {"num":15,"hsn_sc":"38089910","desc":"","user_desc":"","uqc":"PCS","qty":1,"rt":18,"txval":242.55,"iamt":0,"camt":21.83,"samt":21.83,"csamt":0},
        {"num":16,"hsn_sc":"380892","desc":"","user_desc":"","uqc":"PCS","qty":2,"rt":18,"txval":138.6,"iamt":0,"camt":12.47,"samt":12.47,"csamt":0},
    ],
    "03-2026": [
        {"num":1, "hsn_sc":"31059010","desc":"","user_desc":"","uqc":"PCS","qty":79,  "rt":5, "txval":146634.28,"iamt":0,"camt":3665.86,"samt":3665.86,"csamt":0},
        {"num":2, "hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PCS","qty":19,  "rt":5, "txval":120982.86,"iamt":0,"camt":3024.57,"samt":3024.57,"csamt":0},
        {"num":3, "hsn_sc":"38089349","desc":"","user_desc":"","uqc":"PAC","qty":79,  "rt":5, "txval":115239.99,"iamt":0,"camt":2881.00,"samt":2881.00,"csamt":0},
        {"num":4, "hsn_sc":"38089290","desc":"","user_desc":"","uqc":"PCS","qty":2566,"rt":18,"txval":629572.88,"iamt":0,"camt":56661.56,"samt":56661.56,"csamt":0},
        {"num":5, "hsn_sc":"38089390","desc":"","user_desc":"","uqc":"PCS","qty":619, "rt":18,"txval":102203.39,"iamt":0,"camt":9198.31,"samt":9198.31,"csamt":0},
        {"num":6, "hsn_sc":"38089390","desc":"","user_desc":"","uqc":"PAC","qty":206, "rt":18,"txval":54167.80, "iamt":0,"camt":4875.10,"samt":4875.10,"csamt":0},
        {"num":7, "hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PCS","qty":354, "rt":18,"txval":88916.95, "iamt":0,"camt":8002.53,"samt":8002.53,"csamt":0},
        {"num":8, "hsn_sc":"38089199","desc":"","user_desc":"","uqc":"PAC","qty":29,  "rt":18,"txval":29638.98, "iamt":0,"camt":2667.51,"samt":2667.51,"csamt":0},
        {"num":9, "hsn_sc":"39201012","desc":"","user_desc":"","uqc":"PCS","qty":59,  "rt":18,"txval":47013.56, "iamt":0,"camt":4231.22,"samt":4231.22,"csamt":0},
        {"num":10,"hsn_sc":"38089350","desc":"","user_desc":"","uqc":"PCS","qty":88,  "rt":18,"txval":24528.81, "iamt":0,"camt":2207.59,"samt":2207.59,"csamt":0},
        {"num":11,"hsn_sc":"34029099","desc":"","user_desc":"","uqc":"PCS","qty":206, "rt":18,"txval":21462.71, "iamt":0,"camt":1931.64,"samt":1931.64,"csamt":0},
        {"num":12,"hsn_sc":"3808",    "desc":"","user_desc":"","uqc":"PCS","qty":59,  "rt":18,"txval":14308.47, "iamt":0,"camt":1287.76,"samt":1287.76,"csamt":0},
        {"num":13,"hsn_sc":"38089910","desc":"","user_desc":"","uqc":"PCS","qty":29,  "rt":18,"txval":7154.24,  "iamt":0,"camt":643.88, "samt":643.88, "csamt":0},
        {"num":14,"hsn_sc":"380892",  "desc":"","user_desc":"","uqc":"PCS","qty":44,  "rt":18,"txval":3066.11,  "iamt":0,"camt":275.95, "samt":275.95, "csamt":0},
    ],
}

# ── Supplier master ───────────────────────────────────────────────────────────
DEFAULT_SUPPLIERS = {
    "Mandavya AgroChem Pvt Ltd":    "29AALCM6939N1ZG",
    "Sri Manjunatha Agro Agencies": "29ABAFS1520L1Z6",
    "Mahalakshmi Agro Chemicals":   "29AJMPR7424E1ZT",
    "Rallis India Limited KA":      "29AABCR2657NIZU",
    "Rallis India Limited TN":      "33AABCR2657NIZ5",
}

# Purchase → Sales conversion ratios (from Oct-2025 actuals)
RATIO_18_TO_18 = 0.912
RATIO_18_TO_12 = 0.0064
RATIO_5_TO_5   = 1.25
RATIO_EX_TO_EX = 0.865

# HSN proportions from Oct-2025 GSTR-1


# ── Pre-loaded B2CS and NIL data from filed GSTR-1 JSONs (Oct–Feb) ───────────
# These are the EXACT values the portal accepted. For already-filed months,
# make_gstr1_json uses these directly so HSN and B2CS always match.
FILED_B2CS = {
    "10-2025": [
        {"typ":"OE","sply_ty":"INTRA","rt":5, "pos":"29","txval":10298.03,"camt":257.45, "samt":257.45, "csamt":0},
        {"typ":"OE","sply_ty":"INTRA","rt":12,"pos":"29","txval":2175,    "camt":130.50, "samt":130.50, "csamt":0},
        {"typ":"OE","sply_ty":"INTRA","rt":18,"pos":"29","txval":339490,  "camt":30554.10,"samt":30554.10,"csamt":0},
    ],
    "11-2025": [
        {"typ":"OE","sply_ty":"INTRA","rt":5, "pos":"29","txval":4830.599999999999,"camt":120.76,"samt":120.76,"csamt":0},
        {"typ":"OE","sply_ty":"INTRA","rt":12,"pos":"29","txval":1374.5,           "camt":82.47, "samt":82.47, "csamt":0},
        {"typ":"OE","sply_ty":"INTRA","rt":18,"pos":"29","txval":274300.0,         "camt":24687.0,"samt":24687.0,"csamt":0},
    ],
    "12-2025": [
        {"typ":"OE","sply_ty":"INTRA","rt":5, "pos":"29","txval":57400.97999999999,"camt":1435.02,"samt":1435.02,"csamt":0},
        {"typ":"OE","sply_ty":"INTRA","rt":12,"pos":"29","txval":281.02,           "camt":16.86,  "samt":16.86,  "csamt":0},
        {"typ":"OE","sply_ty":"INTRA","rt":18,"pos":"29","txval":112650.0,         "camt":10138.5,"samt":10138.5,"csamt":0},
    ],
    "01-2026": [
        {"typ":"OE","sply_ty":"INTRA","rt":5, "pos":"29","txval":56423.25,            "camt":1410.58,"samt":1410.58,"csamt":0},
        {"typ":"OE","sply_ty":"INTRA","rt":12,"pos":"29","txval":798.7,               "camt":47.92,  "samt":47.92,  "csamt":0},
        {"typ":"OE","sply_ty":"INTRA","rt":18,"pos":"29","txval":398700.00000000006,  "camt":35883.0,"samt":35883.0,"csamt":0},
    ],
    "02-2026": [
        {"typ":"OE","sply_ty":"INTRA","rt":5, "pos":"29","txval":144559.73999999996, "camt":3613.99,"samt":3613.99,"csamt":0},
        {"typ":"OE","sply_ty":"INTRA","rt":12,"pos":"29","txval":123.72,             "camt":7.42,   "samt":7.42,   "csamt":0},
        {"typ":"OE","sply_ty":"INTRA","rt":18,"pos":"29","txval":34650.0,            "camt":3118.5, "samt":3118.5, "csamt":0},
    ],
    "03-2026": [
        {"typ":"OE","sply_ty":"INTRA","rt":5, "pos":"29","txval":382857.13,"camt":9571.43, "samt":9571.43, "csamt":0},
        {"typ":"OE","sply_ty":"INTRA","rt":18,"pos":"29","txval":1022033.9,"camt":91983.05,"samt":91983.05,"csamt":0},
    ],
}

FILED_NIL = {
    "10-2025": {"inv":[{"sply_ty":"INTRAB2C","nil_amt":0,"expt_amt":11250,  "ngsup_amt":0}]},
    "11-2025": {"inv":[{"sply_ty":"INTRAB2C","nil_amt":0,"expt_amt":65740.0,"ngsup_amt":0}]},
    "12-2025": {"inv":[{"sply_ty":"INTRAB2C","nil_amt":0,"expt_amt":0.0,    "ngsup_amt":0}]},
    "01-2026": {"inv":[{"sply_ty":"INTRAB2C","nil_amt":0,"expt_amt":0.0,    "ngsup_amt":0}]},
    "02-2026": {"inv":[{"sply_ty":"INTRAB2C","nil_amt":0,"expt_amt":6920.0, "ngsup_amt":0}]},
    "03-2026": {"inv":[{"sply_ty":"INTRAB2C","nil_amt":0,"expt_amt":0.0,   "ngsup_amt":0}]},
}

FILED_DOC_ISSUE = {
    "10-2025": {"doc_det":[{"doc_num":1,"docs":[{"cancel":0,"from":"351","net_issue":124,"num":1,"to":"474","totnum":124}]}]},
    "11-2025": {"doc_det":[{"doc_num":1,"docs":[{"cancel":0,"from":"475","net_issue":117,"num":1,"to":"591","totnum":117}]}]},
    "12-2025": {"doc_det":[{"doc_num":1,"docs":[{"cancel":0,"from":"592","net_issue":119,"num":1,"to":"710","totnum":119}]}]},
    "01-2026": {"doc_det":[{"doc_num":1,"docs":[{"cancel":0,"from":"714","net_issue":117,"num":1,"to":"830","totnum":117}]}]},
    "02-2026": {"doc_det":[{"doc_num":1,"docs":[{"cancel":0,"from":"834","net_issue":106,"num":1,"to":"939","totnum":106}]}]},
    "03-2026": {"doc_det":[{"doc_num":1,"docs":[{"cancel":0,"from":"940","net_issue":131,"num":1,"to":"1070","totnum":131}]}]},
}

# ─────────────────────────────────────────────────────────────────────────────
# DATABASE  (gst_data.db — persists bills, overrides, suppliers across sessions)
# ─────────────────────────────────────────────────────────────────────────────
DB_FILE = "gst_data.db"

def db_conn():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def db_init():
    """Create tables on first run."""
    with db_conn() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS bills (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                month_key TEXT    NOT NULL,
                data_json TEXT    NOT NULL
            );
            CREATE TABLE IF NOT EXISTS sales_overrides (
                month_key TEXT NOT NULL,
                day       INTEGER NOT NULL,
                data_json TEXT NOT NULL,
                PRIMARY KEY (month_key, day)
            );
            CREATE TABLE IF NOT EXISTS suppliers (
                name  TEXT PRIMARY KEY,
                gstin TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS hsn_entries (
                month_key TEXT    NOT NULL,
                data_json TEXT    NOT NULL
            );
            CREATE TABLE IF NOT EXISTS monthly_targets (
                month_key     TEXT PRIMARY KEY,
                target_amount REAL NOT NULL
            );
        """)

def db_load_bills(mk):
    with db_conn() as conn:
        rows = conn.execute("SELECT data_json FROM bills WHERE month_key=? ORDER BY id", (mk,)).fetchall()
    return [json.loads(r["data_json"]) for r in rows]

def db_save_bills(mk, bills_list):
    """Replace all bills for a month."""
    with db_conn() as conn:
        conn.execute("DELETE FROM bills WHERE month_key=?", (mk,))
        for b in bills_list:
            safe = {k: v for k, v in b.items() if not isinstance(v, bytes)}
            # Convert date objects to strings for JSON serialisation
            for fld in ("inv_date_obj",):
                if fld in safe and hasattr(safe[fld], "isoformat"):
                    safe[fld] = safe[fld].isoformat()
            conn.execute("INSERT INTO bills (month_key, data_json) VALUES (?,?)",
                         (mk, json.dumps(safe)))

def db_load_overrides(mk):
    with db_conn() as conn:
        rows = conn.execute("SELECT day, data_json FROM sales_overrides WHERE month_key=?", (mk,)).fetchall()
    return {r["day"]: json.loads(r["data_json"]) for r in rows}

def db_save_override(mk, day, data):
    with db_conn() as conn:
        conn.execute("INSERT OR REPLACE INTO sales_overrides (month_key, day, data_json) VALUES (?,?,?)",
                     (mk, day, json.dumps(data)))

def db_delete_overrides(mk):
    with db_conn() as conn:
        conn.execute("DELETE FROM sales_overrides WHERE month_key=?", (mk,))

def db_load_target(mk):
    """Return saved target_amount for month_key, or None."""
    with db_conn() as conn:
        row = conn.execute(
            "SELECT target_amount FROM monthly_targets WHERE month_key=?", (mk,)
        ).fetchone()
    return row["target_amount"] if row else None

def db_save_target(mk, amount):
    """Upsert a target_amount for month_key."""
    with db_conn() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO monthly_targets (month_key, target_amount) VALUES (?,?)",
            (mk, amount)
        )

def db_load_suppliers():
    with db_conn() as conn:
        rows = conn.execute("SELECT name, gstin FROM suppliers").fetchall()
    return {r["name"]: r["gstin"] for r in rows}

def db_save_supplier(name, gstin):
    with db_conn() as conn:
        conn.execute("INSERT OR REPLACE INTO suppliers (name, gstin) VALUES (?,?)", (name, gstin))

def db_load_hsn(mk):
    """Load HSN entries for a month from DB. Returns list of dicts."""
    with db_conn() as conn:
        row = conn.execute(
            "SELECT data_json FROM hsn_entries WHERE month_key=?", (mk,)
        ).fetchone()
    return json.loads(row["data_json"]) if row else None

def db_save_hsn(mk, entries):
    """Persist HSN entries for a month."""
    with db_conn() as conn:
        conn.execute("DELETE FROM hsn_entries WHERE month_key=?", (mk,))
        conn.execute(
            "INSERT INTO hsn_entries (month_key, data_json) VALUES (?,?)",
            (mk, json.dumps(entries))
        )

# Initialise DB on startup
db_init()

def ss_init():
    if "page" not in st.session_state:
        st.session_state.page = "dashboard"
    if "gemini_key" not in st.session_state:
        st.session_state.gemini_key = ""
    # Load bills from DB (runs once per session)
    if "bills" not in st.session_state:
        st.session_state.bills = {}
        for mk, _ in MONTHS:
            st.session_state.bills[mk] = db_load_bills(mk)
    # Load overrides from DB
    if "sales_override" not in st.session_state:
        st.session_state.sales_override = {}
        for mk, _ in MONTHS:
            st.session_state.sales_override[mk] = db_load_overrides(mk)
    # Load suppliers: DB takes priority over defaults, merge both
    if "suppliers" not in st.session_state:
        merged = dict(DEFAULT_SUPPLIERS)
        merged.update(db_load_suppliers())
        st.session_state.suppliers = merged
    # Load HSN entries: seed from FILED_HSN if not yet in DB
    if "hsn_entries" not in st.session_state:
        st.session_state.hsn_entries = {}
        for mk, _ in MONTHS:
            from_db = db_load_hsn(mk)
            if from_db is not None:
                st.session_state.hsn_entries[mk] = from_db
            elif mk in FILED_HSN and FILED_HSN[mk]:
                # Seed from filed data — save to DB so edits persist
                st.session_state.hsn_entries[mk] = list(FILED_HSN[mk])
                db_save_hsn(mk, FILED_HSN[mk])
            else:
                st.session_state.hsn_entries[mk] = []
ss_init()
# ─────────────────────────────────────────────────────────────────────────────
# SUPPLIER HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def get_supplier_names():
    """Return sorted list of known supplier names."""
    return sorted(st.session_state.suppliers.keys())

def get_gstin_for(name):
    """Return GSTIN for a supplier name, or empty string."""
    return st.session_state.suppliers.get(name, "")

def add_supplier(name, gstin):
    """Persist a new supplier to session state and DB."""
    name = name.strip(); gstin = gstin.strip()
    if name and gstin:
        st.session_state.suppliers[name] = gstin
        db_save_supplier(name, gstin)

def supplier_selectbox(key_prefix, current_name="", current_gstin=""):
    """Plain text inputs — type supplier name and GSTIN freely."""
    nn_key = f"{key_prefix}_ntxt"
    ng_key = f"{key_prefix}_gtxt"
    if nn_key not in st.session_state:
        st.session_state[nn_key] = current_name
    if ng_key not in st.session_state:
        st.session_state[ng_key] = current_gstin
    ca, cb = st.columns(2)
    ca.text_input("Supplier name",  key=nn_key, placeholder="e.g. Rallis India Ltd")
    cb.text_input("Supplier GSTIN", key=ng_key, placeholder="e.g. 29AABCR2657NIZU")
    return st.session_state[nn_key].strip(), st.session_state[ng_key].strip()

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📋 SLVT GST Tool")
    st.markdown("---")
    if st.button("🏠  Dashboard"):
        st.session_state.page = "dashboard"
    st.markdown("**Monthly registers**")
    for mk, lbl in MONTHS:
        bs_count = len([b for b in st.session_state.bills.get(mk,[]) if b.get("status")=="ok"])
        icon = "✅" if bs_count > 0 else "⏳"
        if st.button(f"  {icon} {lbl}", key=f"nav_{mk}"):
            st.session_state.page = mk
    st.markdown("---")
    st.markdown("**Gemini API Key**")
    st.caption("For AI bill photo scanning")
    key_in = st.text_input("", value=st.session_state.gemini_key,
                            type="password", placeholder="AIza…",
                            label_visibility="collapsed")
    if key_in:
        st.session_state.gemini_key = key_in
    if st.session_state.gemini_key:
        st.success("✓ Key set")
    else:
        st.caption("Get free key: aistudio.google.com")
    st.markdown("---")
    st.caption(f"GSTIN: {GSTIN}")

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def r2(n):   return round((n or 0)*100)/100
def fmtc(n): return f"₹{r2(n):,.2f}"

def parse_mk(mk):
    m, y = mk.split("-"); return int(m), int(y)

def ld(m, y): return calendar.monthrange(y, m)[1]

def ok_bills(mk):
    return [b for b in st.session_state.bills.get(mk,[]) if b.get("status")=="ok"]

def bills_summary(mk):
    bills = ok_bills(mk)
    p18  = sum(b.get("val18") or 0 for b in bills)
    p5   = sum(b.get("val5")  or 0 for b in bills)
    p12  = sum(b.get("val12") or 0 for b in bills)
    pex  = sum(b.get("exempt")or 0 for b in bills)
    itc  = sum((b.get("cgst9") or 0)+(b.get("sgst9") or 0)+
               (b.get("cgst25")or 0)+(b.get("sgst25")or 0) for b in bills)
    gross= sum(b.get("gross") or 0 for b in bills)
    return {"p18":r2(p18),"p5":r2(p5),"p12":r2(p12),"pex":r2(pex),
            "itc":r2(itc),"gross":r2(gross),"count":len(bills)}

def derive_sales_totals(mk):
    bs = bills_summary(mk)
    return {
        "v18": r2(bs["p18"]*RATIO_18_TO_18),
        "v5":  r2(bs["p5"] *RATIO_5_TO_5),
        "v12": r2(bs["p18"]*RATIO_18_TO_12),
        "vex": r2(bs["pex"]*RATIO_EX_TO_EX),
    }

def get_voucher_start(mk):
    """Return voucher start number from filed GSTR-1 JSON lookup table."""
    return VOUCHER_START.get(mk, OCT_LAST_VNO + 1)

def get_purchase_voucher_start(mk):
    """Return next purchase voucher number continuing from Oct (ended at 61)."""
    mk_list = [m for m,_ in MONTHS]
    idx = mk_list.index(mk)
    vno = OCT_LAST_PUR_VNO + 1
    for i in range(idx):
        prev_mk = mk_list[i]
        vno += len(ok_bills(prev_mk))
    return vno

def build_daily_sales(mk):
    m, y    = parse_mk(mk)
    days    = ld(m, y)
    totals  = derive_sales_totals(mk)
    overrides = st.session_state.sales_override.get(mk, {})
    random.seed(m*1000+y)

    all_days = list(range(1, days))
    random.shuffle(all_days)
    only5  = set(all_days[:3])
    mixed  = set(all_days[3:6])
    has12  = set(random.sample(all_days[6:], min(2, max(0,len(all_days)-6))))

    running = {"v18":0.0,"v5":0.0,"v12":0.0}
    entries = []

    for d in range(1, days+1):
        if d in overrides:
            ov = overrides[d]
            entries.append({"date":datetime.date(y,m,d),
                "v18":ov.get("v18",0),"v5":ov.get("v5",0),
                "v12":ov.get("v12",0),"vex":ov.get("vex",0),
                "nbills":ov.get("nbills",4)})
            for k in ("v18","v5","v12"):
                running[k] += ov.get(k,0)
            continue

        is_last = (d == days)
        if is_last:
            vex = max(0, r2(totals["vex"] - sum(
                e["vex"] for e in entries)))
            entries.append({"date":datetime.date(y,m,d),
                "v18":0,"v5":0,"v12":0,"vex":vex,
                "nbills":random.randint(2,4)})
            continue

        rem_days  = days - d
        rem18 = max(0, totals["v18"] - running["v18"])
        rem5  = max(0, totals["v5"]  - running["v5"])
        rem12 = max(0, totals["v12"] - running["v12"])
        avg18 = rem18 / max(rem_days,1)
        avg5  = rem5  / max(rem_days,1)
        avg12 = rem12 / max(rem_days,1)
        dm    = min(max(random.lognormvariate(0, 0.75), 0.08), 3.5)

        if d in only5:
            v18=0; v5=r2(min(avg5*random.uniform(2,3.5), rem5)); v12=0; nb=random.randint(2,4)
        elif d in mixed:
            v18=r2(min(avg18*dm*0.8, rem18)); v5=r2(min(avg5*random.uniform(1.5,3), rem5)); v12=0; nb=random.randint(3,6)
        elif d in has12:
            v18=r2(min(avg18*dm*0.95, rem18)); v12=r2(min(avg12*random.uniform(1.5,2.5), rem12)); v5=0; nb=random.randint(4,7)
        else:
            v18=r2(min(avg18*dm, rem18)); v5=0; v12=0; nb=random.randint(2,6)

        if v18 > 0:
            v18 = float(round(v18/50)*50)
        v18=max(0,v18); v5=max(0,v5); v12=max(0,v12)

        entries.append({"date":datetime.date(y,m,d),"v18":v18,"v5":v5,"v12":v12,"vex":0,"nbills":nb})
        running["v18"]+=v18; running["v5"]+=v5; running["v12"]+=v12

    return entries

def distribute_target_sales(mk, target_val):
    """
    Smart Sales Override: clears all existing overrides for `mk` and generates
    new daily entries whose grand total (gross) equals exactly `target_val`.

    Tax split — matches real SLVT sales pattern (no exempt on regular days):
        75% of gross revenue → 18% GST slab  (taxable = portion / 1.18)
        25% of gross revenue → 5%  GST slab  (taxable = portion / 1.05)
        Exempt = 0  (exempt items are rare one-off entries, not daily)

    Daily distribution uses ±30% log-normal variation so no day looks flat.
    The final day absorbs any rounding remainder to guarantee exact totals.
    Voucher count per day: 2–6 (mirrors the Feb register pattern).
    """
    m, y  = parse_mk(mk)
    days  = ld(m, y)

    # ── Split target gross into GST slabs ──────────────────────────────────
    # gross = taxable * 1.18  (for 18% slab)  or  taxable * 1.05  (for 5%)
    rev18 = r2(target_val * 0.75)          # 75 % of gross through 18% slab
    rev5  = r2(target_val - rev18)         # remaining 25 % through 5% slab

    # Convert gross revenue to taxable value
    total18 = r2(rev18 / 1.18)
    total5  = r2(rev5  / 1.05)

    # ── Generate per-day weights with ±30 % lognormal variation ─────────────
    random.seed(mk + "_smart")
    raw_weights = [max(0.4, min(random.lognormvariate(0, 0.30), 1.8))
                   for _ in range(days)]
    wsum = sum(raw_weights)
    weights = [w / wsum for w in raw_weights]

    # ── First clear existing overrides, then build new ones ─────────────────
    db_delete_overrides(mk)
    st.session_state.sales_override[mk] = {}

    running18 = 0.0
    running5  = 0.0
    running_gross = 0.0

    for d in range(1, days + 1):
        w       = weights[d - 1]
        is_last = (d == days)

        if is_last:
            # Absorb all remaining so gross sums exactly to target_val
            v18 = r2(max(0, total18 - running18))
            v5  = r2(max(0, total5  - running5))
            # Fine-tune v5 so gross lands exactly on target
            expected_gross = r2(v18 * 1.18 + v5 * 1.05)
            remainder = r2(target_val - running_gross - expected_gross)
            if remainder != 0:
                v5 = r2(v5 + remainder / 1.05)
        else:
            v18 = r2(total18 * w)
            v5  = r2(total5  * w)

        nbills  = random.randint(2, 6)
        ov_data = {"v18": v18, "v5": v5, "v12": 0.0, "vex": 0.0, "nbills": nbills}

        st.session_state.sales_override[mk][d] = ov_data
        db_save_override(mk, d, ov_data)

        running18 += v18
        running5  += v5
        running_gross += r2(v18 * 1.18 + v5 * 1.05)

    db_save_target(mk, target_val)


def split_evenly(total, n):
    if not total or n<1: return [0.0]*max(n,1)
    base=r2(total/n); parts=[base]*n; parts[-1]=r2(total-base*(n-1))
    return parts

# ─────────────────────────────────────────────────────────────────────────────
# XLSX GENERATORS
# ─────────────────────────────────────────────────────────────────────────────
def header_block(reg_type, m, y):
    ys=str(y)[2:]; yn=str(y+1)[2:]; abbr=MON_ABBR[m]; last=ld(m,y)
    return [[f"{SHOP_NAME}-{ys}-{yn}"],[SHOP_ADDR1],[SHOP_ADDR2],
            [SHOP_PHONE],[SHOP_STATE],[reg_type],
            [f"1-{abbr}-{ys} to {last}-{abbr}-{ys}"]]

def set_widths(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

def to_bytes(wb):
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

def make_sales_xlsx(mk, entries, start_vno):
    m,y=parse_mk(mk)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Sales Register"
    for row in header_block("Sales Register",m,y): ws.append(row)
    ws.append(["Date","Particulars","Buyer","Voucher Type","Voucher No.",
        "Voucher Ref. No.","GSTIN/UIN","Value","Gross Total",
        "18% Sales","9% CGST Output","9% SGST  Output",
        "5% Sales","2.5% Cgst Output","2.5% Sgst Output",
        "12% Sales","6% CGST Output","6% SGCT Output","Exampted Sales"])
    vno=start_vno
    g={k:0.0 for k in ["val","gross","v18","c9","s9","v5","c25","s25","v12","c6","s6","ex"]}
    for e in entries:
        b18=split_evenly(e["v18"],e["nbills"]); b5=split_evenly(e["v5"],e["nbills"])
        b12=split_evenly(e["v12"],e["nbills"]); bex=split_evenly(e["vex"],e["nbills"])
        for i in range(e["nbills"]):
            bv18=b18[i];bv5=b5[i];bv12=b12[i];bvex=bex[i]
            bval=r2(bv18+bv5+bv12+bvex)
            if bval==0: continue
            bgross=r2(bv18*1.18+bv5*1.05+bv12*1.12+bvex)
            ws.append([e["date"],"Cash","Cash","Sales",str(vno),None,None,
                bval,bgross,
                bv18 or None,r2(bv18*.09) or None,r2(bv18*.09) or None,
                bv5  or None,r2(bv5*.025) or None,r2(bv5*.025) or None,
                bv12 or None,r2(bv12*.06) or None,r2(bv12*.06) or None,
                bvex or None])
            g["val"]+=bval;g["gross"]+=bgross
            g["v18"]+=bv18;g["c9"]+=r2(bv18*.09);g["s9"]+=r2(bv18*.09)
            g["v5"]+=bv5;g["c25"]+=r2(bv5*.025);g["s25"]+=r2(bv5*.025)
            g["v12"]+=bv12;g["c6"]+=r2(bv12*.06);g["s6"]+=r2(bv12*.06)
            g["ex"]+=bvex; vno+=1
    ws.append([None,"Grand Total",None,None,None,None,None,
        r2(g["val"]),r2(g["gross"]),
        r2(g["v18"]) or None,r2(g["c9"]) or None,r2(g["s9"]) or None,
        r2(g["v5"])  or None,r2(g["c25"])or None,r2(g["s25"])or None,
        r2(g["v12"]) or None,r2(g["c6"]) or None,r2(g["s6"]) or None,
        r2(g["ex"])  or None])
    set_widths(ws,[12,14,12,12,10,14,16,10,10,10,13,13,10,14,14,10,13,13,13])
    return to_bytes(wb), g, vno-1

def make_purchase_xlsx(mk, purchases):
    m,y=parse_mk(mk)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Purchase Register"
    for row in header_block("Purchase Register",m,y): ws.append(row)
    ws.append(["Date","Particulars","Supplier","Voucher Type","Voucher No.",
        "Supplier Invoice No.","Supplier Invoice Date","GSTIN/UIN",
        "Value","Addl. Cost","Gross Total",
        "18 % Purchase","9% CGST","9% SGST","Round Off",
        "5 % Purchase","2.5% CGST","2.5 % SGST","EXAMPTED","CASH DISCOUNT"])
    g={k:0.0 for k in ["val","gross","v18","c9","s9","ro","v5","c25","s25","ex","disc"]}
    for i,p in enumerate(purchases,1):
        val=r2((p.get("val18") or 0)+(p.get("val5") or 0)+(p.get("val12") or 0)+(p.get("exempt") or 0))
        ws.append([p.get("inv_date_obj") or p.get("inv_date",""),
            p.get("supplier",""),p.get("supplier",""),"Purchase",str(p.get("pur_vno", i)),
            p.get("invno",""),p.get("inv_date_obj") or p.get("inv_date",""),p.get("gstin",""),
            val,None,r2(p.get("gross") or 0),
            p.get("val18") or None,p.get("cgst9") or None,p.get("sgst9") or None,p.get("round_off") or None,
            p.get("val5")  or None,p.get("cgst25")or None,p.get("sgst25")or None,
            p.get("exempt")or None,p.get("discount")or None])
        g["val"]+=val;g["gross"]+=p.get("gross") or 0
        g["v18"]+=p.get("val18") or 0;g["c9"]+=p.get("cgst9") or 0;g["s9"]+=p.get("sgst9") or 0
        g["v5"]+=p.get("val5") or 0;g["c25"]+=p.get("cgst25") or 0;g["s25"]+=p.get("sgst25") or 0
        g["ex"]+=p.get("exempt") or 0;g["ro"]+=p.get("round_off") or 0;g["disc"]+=p.get("discount") or 0
    ws.append([None,"Grand Total",None,None,None,None,None,None,
        r2(g["val"]),None,r2(g["gross"]),
        r2(g["v18"]) or None,r2(g["c9"]) or None,r2(g["s9"]) or None,r2(g["ro"]) or None,
        r2(g["v5"])  or None,r2(g["c25"])or None,r2(g["s25"])or None,
        r2(g["ex"])  or None,r2(g["disc"])or None])
    set_widths(ws,[12,20,20,10,10,18,16,18,12,8,12,14,10,10,10,12,10,10,10,12])
    return to_bytes(wb), g

# ─────────────────────────────────────────────────────────────────────────────
# GSTR-1 JSON
# ─────────────────────────────────────────────────────────────────────────────
def make_gstr1_json(mk, sales_g, vno_from, vno_to, vno_count):
    fp = MONTH_FP.get(mk, mk.replace("-",""))

    # ── For already-filed months (Nov–Feb): use EXACT filed values ────────────
    # This guarantees b2cs, nil, doc_issue and HSN are all internally consistent
    # and match what the portal previously accepted — preventing schema rejection.
    if mk in FILED_B2CS:
        b2cs      = FILED_B2CS[mk]
        nil       = FILED_NIL[mk]
        doc_issue = FILED_DOC_ISSUE[mk]
    else:
        # ── March and future months: compute from sales register ──────────────
        v18=sales_g["v18"]; v5=sales_g["v5"]; v12=sales_g["v12"]; vex=sales_g["ex"]
        b2cs=[]
        for rt,txval in [(5,v5),(12,v12),(18,v18)]:
            if txval>0:
                rf=rt/100/2
                b2cs.append({"typ":"OE","sply_ty":"INTRA","rt":rt,"pos":"29",
                    "txval":r2(txval),"camt":r2(txval*rf),"samt":r2(txval*rf),"csamt":0})
        nil={"inv":[{"sply_ty":"INTRAB2C","nil_amt":0,"expt_amt":vex,"ngsup_amt":0}]}
        doc_issue={"doc_det":[{"doc_num":1,"docs":[{
            "cancel":0,"from":str(vno_from),"net_issue":vno_count,
            "num":1,"to":str(vno_to),"totnum":vno_count}]}]}

    result = {"gstin":GSTIN,"fp":fp,"b2cs":b2cs,"nil":nil,"doc_issue":doc_issue}

    # ── HSN section ───────────────────────────────────────────────────────────
    # Key order matches portal-accepted format from Oct 2025 accepted GSTR-1:
    # gstin → fp → b2cs → nil → hsn → doc_issue  (hsn MUST come before doc_issue)
    HSN_KEY_ORDER = ["num","hsn_sc","txval","iamt","camt","samt","csamt","desc","user_desc","uqc","qty","rt"]
    hsn_rows = st.session_state.get("hsn_entries", {}).get(mk, [])
    if hsn_rows:
        ordered_rows = [
            {k: row.get(k, 0 if k not in ("desc","user_desc","hsn_sc","uqc") else "")
             for k in HSN_KEY_ORDER}
            for row in hsn_rows
        ]
        # Rebuild result with hsn inserted before doc_issue
        result = {"gstin":GSTIN,"fp":fp,"b2cs":b2cs,"nil":nil,
                  "hsn":{"hsn_b2c":ordered_rows},"doc_issue":doc_issue}

    return result

# ─────────────────────────────────────────────────────────────────────────────
# AI BILL EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────
EXTRACT_PROMPT = """Extract all fields from this Indian GST purchase invoice.
Return ONLY valid JSON — no markdown, no explanation.
{
  "supplier":"full company name","gstin":"15-char GSTIN",
  "invno":"invoice number","inv_date":"DD-MM-YYYY",
  "val18":taxable@18% (number),"cgst9":CGST9% (number),"sgst9":SGST9% (number),
  "val5":taxable@5% (number),"cgst25":CGST2.5% (number),"sgst25":SGST2.5% (number),
  "val12":taxable@12% (number),"exempt":exempt value (number),
  "gross":total invoice amount (number),"round_off":rounding (number),"discount":cash discount (number)
}
All numbers plain — no commas, no symbols. Use 0 if not present."""

def extract_bill_ai(image_bytes, mime_type, api_key):
    b64=base64.b64encode(image_bytes).decode()
    resp=requests.post(
        f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}",
        headers={"content-type":"application/json"},
        json={"contents":[{"parts":[{"inline_data":{"mime_type":mime_type,"data":b64}},
            {"text":EXTRACT_PROMPT}]}],
            "generationConfig":{"maxOutputTokens":500,"temperature":0}},
        timeout=30)
    resp.raise_for_status()
    text=resp.json()["candidates"][0]["content"]["parts"][0]["text"]
    text=re.sub(r"```[a-z]*","",text).strip()
    data=json.loads(text)
    try:
        pts=str(data.get("inv_date","")).split("-")
        data["inv_date_obj"]=datetime.date(int(pts[2]),int(pts[1]),int(pts[0]))
    except: data["inv_date_obj"]=None
    for f in ["val18","cgst9","sgst9","val5","cgst25","sgst25","val12",
              "exempt","gross","round_off","discount"]:
        try:    data[f]=float(data.get(f) or 0)
        except: data[f]=0.0
    return data

# ─────────────────────────────────────────────────────────────────────────────
# SHARED BILL FIELDS COMPONENT
# Returns a dict of updated values — caller must write back to session state
# ─────────────────────────────────────────────────────────────────────────────
def bill_fields(bill, kp):
    """Render editable fields for a bill. Returns updated dict."""
    updated = dict(bill)  # work on a copy

    # ── Supplier + invoice header ─────────────────────────────────────────────
    c_sup, c_inv, c_dt = st.columns([2,1,1])
    with c_sup:
        updated["supplier"], updated["gstin"] = supplier_selectbox(
            f"{kp}s", current_name=bill.get("supplier",""), current_gstin=bill.get("gstin",""))
    updated["invno"]   = c_inv.text_input("Supplier Invoice no.", value=bill.get("invno",""),    key=f"{kp}inv")
    updated["inv_date"]= c_dt.text_input( "Date DD-MM-YYYY",      value=bill.get("inv_date",""), key=f"{kp}dt")
    updated["pur_vno"] = st.number_input("Voucher No. (Tally)",
                             min_value=1, value=int(bill.get("pur_vno") or 1),
                             step=1, key=f"{kp}pvno")

    # ── GST slabs — taxable auto-fills GST, but all fields editable ─────────
    _def = bill  # stored values take priority

    st.markdown("<div style='font-size:11px;font-weight:600;color:#4a7a47;"
        "text-transform:uppercase;letter-spacing:.4px;margin:8px 0 3px'>"
        "18% slab</div>", unsafe_allow_html=True)
    ea1,ea2,ea3 = st.columns(3)
    updated["val18"] = ea1.number_input("Taxable value", value=float(_def.get("val18") or 0), min_value=0.0, step=1.0, format="%.2f", key=f"{kp}v18")
    updated["cgst9"] = ea2.number_input("CGST 9%",       value=float(_def.get("cgst9") or r2(updated["val18"]*0.09)), min_value=0.0, step=0.01, format="%.2f", key=f"{kp}c9")
    updated["sgst9"] = ea3.number_input("SGST 9%",       value=float(_def.get("sgst9") or r2(updated["val18"]*0.09)), min_value=0.0, step=0.01, format="%.2f", key=f"{kp}s9")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#4a7a47;"
        "text-transform:uppercase;letter-spacing:.4px;margin:8px 0 3px'>"
        "5% slab</div>", unsafe_allow_html=True)
    eb1,eb2,eb3 = st.columns(3)
    updated["val5"]   = eb1.number_input("Taxable value", value=float(_def.get("val5")   or 0), min_value=0.0, step=1.0,  format="%.2f", key=f"{kp}v5")
    updated["cgst25"] = eb2.number_input("CGST 2.5%",     value=float(_def.get("cgst25") or r2(updated["val5"]*0.025)), min_value=0.0, step=0.01, format="%.2f", key=f"{kp}c25")
    updated["sgst25"] = eb3.number_input("SGST 2.5%",     value=float(_def.get("sgst25") or r2(updated["val5"]*0.025)), min_value=0.0, step=0.01, format="%.2f", key=f"{kp}s25")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#4a7a47;"
        "text-transform:uppercase;letter-spacing:.4px;margin:8px 0 3px'>"
        "12% slab</div>", unsafe_allow_html=True)
    ec1,ec2,ec3 = st.columns(3)
    updated["val12"] = ec1.number_input("Taxable value", value=float(_def.get("val12") or 0), min_value=0.0, step=1.0,  format="%.2f", key=f"{kp}v12")
    updated["cgst6"] = ec2.number_input("CGST 6%",       value=float(_def.get("cgst6") or r2(updated["val12"]*0.06)), min_value=0.0, step=0.01, format="%.2f", key=f"{kp}c6")
    updated["sgst6"] = ec3.number_input("SGST 6%",       value=float(_def.get("sgst6") or r2(updated["val12"]*0.06)), min_value=0.0, step=0.01, format="%.2f", key=f"{kp}s6")

    st.markdown("<div style='font-size:11px;font-weight:600;color:#4a7a47;"
        "text-transform:uppercase;letter-spacing:.4px;margin:8px 0 3px'>"
        "Other</div>", unsafe_allow_html=True)
    ed1,ed2,ed3,ed4 = st.columns(4)
    updated["exempt"]    = ed1.number_input("Exempt/Nil",    value=float(_def.get("exempt")    or 0), min_value=0.0, step=1.0,  format="%.2f", key=f"{kp}ex")
    updated["round_off"] = ed2.number_input("Round off",     value=float(_def.get("round_off") or 0), step=0.01,               format="%.2f", key=f"{kp}ro")
    updated["discount"]  = ed3.number_input("Cash discount", value=float(_def.get("discount")  or 0), min_value=0.0, step=1.0,  format="%.2f", key=f"{kp}dc")

    total_gst = r2(updated["cgst9"]+updated["sgst9"]+updated["cgst25"]+updated["sgst25"]+updated["cgst6"]+updated["sgst6"])
    auto_gross = r2(updated["val18"]+updated["val5"]+updated["val12"]+updated["exempt"]+total_gst+updated["round_off"]-updated["discount"])
    updated["gross"] = ed4.number_input("Gross total", value=float(_def.get("gross") or auto_gross), min_value=0.0, step=1.0, format="%.2f", key=f"{kp}gr")

    itc = r2(updated["cgst9"]+updated["sgst9"]+updated["cgst25"]+updated["sgst25"])
    st.markdown(
        f"<div style='background:#eaf2e8;border-radius:5px;padding:7px 12px;"
        f"font-size:12px;color:#1a3c1a;margin-top:6px'>"
        f"Total GST: <strong>{fmtc(total_gst)}</strong> &nbsp;|&nbsp; "
        f"ITC: <strong>{fmtc(itc)}</strong> &nbsp;|&nbsp; "
        f"Gross: <strong>{fmtc(updated['gross'])}</strong>"
        f"</div>", unsafe_allow_html=True)
    # ── Date obj ──────────────────────────────────────────────────────────────
    try:
        pts = updated["inv_date"].split("-")
        updated["inv_date_obj"] = datetime.date(int(pts[2]),int(pts[1]),int(pts[0]))
    except: updated["inv_date_obj"] = None

    return updated

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
def page_dashboard():
    st.markdown("## 🏠 Dashboard")
    st.caption("Enter purchase bills for each month → sales auto-derive → download all 3 documents per month")
    st.markdown("---")

    rows=[]
    for mk,lbl in MONTHS:
        bs=bills_summary(mk)
        st_txt="✅ Ready" if bs["count"]>0 else "⏳ Awaiting bills"
        rows.append({"Month":lbl,"Bills entered":bs["count"],
            "Purchase value":fmtc(bs["p18"]+bs["p5"]+bs["pex"]),
            "ITC available":fmtc(bs["itc"]),
            "Estimated sales":fmtc(sum(derive_sales_totals(mk).values())),
            "Status":st_txt})
    st.dataframe(rows,use_container_width=True,hide_index=True)

    st.markdown("---")
    st.markdown("### Download all documents")
    st.caption("Click a month in the sidebar to enter purchase bills. Once bills are entered, download buttons appear below.")

    any_ready=False
    for mk,lbl in MONTHS:
        bs=bills_summary(mk)
        if bs["count"]==0: continue
        any_ready=True
        st.markdown(f"**{lbl}** — {bs['count']} purchase bills · Est. sales {fmtc(sum(derive_sales_totals(mk).values()))}")
        c1,c2,c3=st.columns(3)
        start_vno=get_voucher_start(mk)
        entries  =build_daily_sales(mk)
        purchases=ok_bills(mk)
        sales_xls,sales_g,last_vno=make_sales_xlsx(mk,entries,start_vno)
        pur_xls,_=make_purchase_xlsx(mk,purchases)
        gstr1=make_gstr1_json(mk,sales_g,start_vno,last_vno,last_vno-start_vno+1)
        m2,y2=parse_mk(mk); mname=lbl.replace(" ","_")
        c1.download_button("⬇ Sales Register .xlsx",data=sales_xls,
            file_name=f"Sales_Register_{mname}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dd_s_{mk}")
        c2.download_button("⬇ Purchase Register .xlsx",data=pur_xls,
            file_name=f"Purchase_Register_{mname}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dd_p_{mk}")
        c3.download_button(f"⬇ GSTR-1 JSON",data=json.dumps(gstr1,indent=2).encode(),
            file_name=f"{GSTIN}_GSTR1_{MON_ABBR[m2]}{y2}.json",
            mime="application/json",key=f"dd_g_{mk}")
        st.markdown("")

    if not any_ready:
        st.info("No months have purchase bills yet. Click a month in the sidebar to get started.")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: MONTH
# ─────────────────────────────────────────────────────────────────────────────
def page_month(mk):
    lbl=dict(MONTHS).get(mk,mk); m,y=parse_mk(mk)
    st.markdown(f"## 📅 {lbl}")
    tab_pur,tab_sales,tab_docs=st.tabs(
        ["📥 Purchase Bills","📊 Sales (derived)","📤 Download Documents"])

    # ── TAB 1: PURCHASE BILLS ─────────────────────────────────────────────────
    with tab_pur:
        st.caption("Enter all purchase bills for this month. Sales registers are auto-generated from these.")

        st.markdown("### ✏️ Add bill manually")

        # ── STEP 1: Input form ────────────────────────────────────────────────
        with st.form(f"addbill_{mk}", clear_on_submit=False):
            nb = {}
            st.markdown("**Supplier**")
            nb["supplier"], nb["gstin"] = supplier_selectbox(f"add{mk}")
            c3f, c4f, c5f = st.columns(3)
            nb["invno"]   = c3f.text_input("Supplier Invoice no. *", placeholder="INV-001",                    key=f"ninv{mk}")
            nb["inv_date"]= c4f.text_input("Date DD-MM-YYYY *",      placeholder=f"10-{str(m).zfill(2)}-{y}", key=f"ndt{mk}")
            _next_pur_vno = get_purchase_voucher_start(mk) + len(ok_bills(mk))
            nb["pur_vno"] = c5f.number_input("Voucher No. (Tally)", min_value=1,
                                              value=_next_pur_vno, step=1, key=f"npvno{mk}")

            # ── 18% slab ─────────────────────────────────────────────────
            st.markdown("<div style='font-size:11px;font-weight:600;color:#4a7a47;"
                "text-transform:uppercase;letter-spacing:.4px;margin:8px 0 3px'>"
                "18% slab</div>", unsafe_allow_html=True)
            fa1,fa2,fa3 = st.columns(3)
            nb["val18"] = fa1.number_input("Taxable value", min_value=0.0,value=0.0,step=1.0,format="%.2f",key=f"nv18{mk}")
            nb["cgst9"] = fa2.number_input("CGST 9%  (auto-fills)",  min_value=0.0,value=r2(nb["val18"]*0.09), step=0.01,format="%.2f",key=f"nc9{mk}")
            nb["sgst9"] = fa3.number_input("SGST 9%  (auto-fills)",  min_value=0.0,value=r2(nb["val18"]*0.09), step=0.01,format="%.2f",key=f"ns9{mk}")

            # ── 5% slab ──────────────────────────────────────────────────────
            st.markdown("<div style='font-size:11px;font-weight:600;color:#4a7a47;"
                "text-transform:uppercase;letter-spacing:.4px;margin:8px 0 3px'>"
                "5% slab</div>", unsafe_allow_html=True)
            fb1,fb2,fb3 = st.columns(3)
            nb["val5"]   = fb1.number_input("Taxable value", min_value=0.0,value=0.0,step=1.0,format="%.2f",key=f"nv5{mk}")
            nb["cgst25"] = fb2.number_input("CGST 2.5% (auto-fills)",min_value=0.0,value=r2(nb["val5"]*0.025),step=0.01,format="%.2f",key=f"nc25{mk}")
            nb["sgst25"] = fb3.number_input("SGST 2.5% (auto-fills)",min_value=0.0,value=r2(nb["val5"]*0.025),step=0.01,format="%.2f",key=f"ns25{mk}")

            # ── 12% slab ─────────────────────────────────────────────────────
            st.markdown("<div style='font-size:11px;font-weight:600;color:#4a7a47;"
                "text-transform:uppercase;letter-spacing:.4px;margin:8px 0 3px'>"
                "12% slab</div>", unsafe_allow_html=True)
            fc1,fc2,fc3 = st.columns(3)
            nb["val12"] = fc1.number_input("Taxable value", min_value=0.0,value=0.0,step=1.0,format="%.2f",key=f"nv12{mk}")
            nb["cgst6"] = fc2.number_input("CGST 6%  (auto-fills)",  min_value=0.0,value=r2(nb["val12"]*0.06),step=0.01,format="%.2f",key=f"nc6{mk}")
            nb["sgst6"] = fc3.number_input("SGST 6%  (auto-fills)",  min_value=0.0,value=r2(nb["val12"]*0.06),step=0.01,format="%.2f",key=f"ns6{mk}")

            # ── Other ─────────────────────────────────────────────────────────
            st.markdown("<div style='font-size:11px;font-weight:600;color:#4a7a47;"
                "text-transform:uppercase;letter-spacing:.4px;margin:8px 0 3px'>"
                "Other</div>", unsafe_allow_html=True)
            fd1,fd2,fd3,fd4 = st.columns(4)
            nb["exempt"]    = fd1.number_input("Exempt/Nil",   min_value=0.0,value=0.0,step=1.0,format="%.2f",key=f"nex{mk}")
            nb["round_off"] = fd2.number_input("Round off",    value=0.0,step=0.01,format="%.2f",key=f"nro{mk}")
            nb["discount"]  = fd3.number_input("Cash discount",min_value=0.0,value=0.0,step=1.0,format="%.2f",key=f"ndc{mk}")
            _auto_gross = r2(nb["val18"]+nb["val5"]+nb["val12"]+nb["exempt"]+nb["cgst9"]+nb["sgst9"]+nb["cgst25"]+nb["sgst25"]+nb["cgst6"]+nb["sgst6"]+nb["round_off"]-nb["discount"])
            nb["gross"]     = fd4.number_input("Gross total",  min_value=0.0,value=_auto_gross,step=1.0,format="%.2f",key=f"ngr{mk}")
            photo = st.file_uploader("Photo (optional)", type=["jpg","jpeg","png","webp"], key=f"nph{mk}")

            calc_clicked = st.form_submit_button("🧮  Calculate — preview before adding",
                                                  use_container_width=True)
            if calc_clicked:
                st.session_state[f"calc_{mk}"] = dict(nb)
                st.session_state[f"calc_photo_{mk}"] = photo.name if photo else None

        # ── STEP 2: Preview + confirm (shown after Calculate is clicked) ──────
        prev = st.session_state.get(f"calc_{mk}")
        if prev:
            # Use amounts exactly as entered — no recompute
            p = dict(prev)
            _tgst = r2((p.get("cgst9") or 0)+(p.get("sgst9") or 0)+
                       (p.get("cgst25") or 0)+(p.get("sgst25") or 0)+
                       (p.get("cgst6") or 0)+(p.get("sgst6") or 0))
            if not p.get("gross"):
                p["gross"] = r2((p.get("val18") or 0)+(p.get("val5") or 0)+
                                (p.get("val12") or 0)+(p.get("exempt") or 0)+
                                _tgst+(p.get("round_off") or 0)-(p.get("discount") or 0))
            _itc = r2((p.get("cgst9") or 0)+(p.get("sgst9") or 0)+
                      (p.get("cgst25") or 0)+(p.get("sgst25") or 0))

            st.markdown("---")
            st.markdown("#### 🧾 Bill preview — verify before adding")

            # Header row
            hc1,hc2,hc3,hc4,hc5 = st.columns(5)
            hc1.markdown(f"**Supplier**<br>{p.get('supplier','—')}", unsafe_allow_html=True)
            hc2.markdown(f"**GSTIN**<br>{p.get('gstin','—')}", unsafe_allow_html=True)
            hc3.markdown(f"**Supplier Invoice no.**<br>{p.get('invno','—')}", unsafe_allow_html=True)
            hc4.markdown(f"**Date**<br>{p.get('inv_date','—')}", unsafe_allow_html=True)
            hc5.markdown(f"**Voucher No.**<br>{p.get('pur_vno','—')}", unsafe_allow_html=True)

            st.markdown("")
            # GST breakdown table
            rows = []
            if p["val18"]: rows.append({"Slab":"18%","Taxable":fmtc(p["val18"]),"CGST":fmtc(p["cgst9"]), "SGST":fmtc(p["sgst9"]), "Total GST":fmtc(p["cgst9"]+p["sgst9"])})
            if p["val5"]:  rows.append({"Slab":"5%", "Taxable":fmtc(p["val5"]), "CGST":fmtc(p["cgst25"]),"SGST":fmtc(p["sgst25"]),"Total GST":fmtc(p["cgst25"]+p["sgst25"])})
            if p["val12"]: rows.append({"Slab":"12%","Taxable":fmtc(p["val12"]),"CGST":fmtc(p["cgst6"]), "SGST":fmtc(p["sgst6"]), "Total GST":fmtc(p["cgst6"]+p["sgst6"])})
            if p["exempt"]:rows.append({"Slab":"Exempt","Taxable":fmtc(p["exempt"]),"CGST":"—","SGST":"—","Total GST":"—"})
            if rows:
                st.dataframe(rows, use_container_width=True, hide_index=True)

            # Summary strip
            sc1,sc2,sc3,sc4,sc5 = st.columns(5)
            sc1.metric("Total taxable", fmtc(p["val18"]+p["val5"]+p["val12"]+p["exempt"]))
            sc2.metric("Total GST",     fmtc(_tgst))
            sc3.metric("ITC",           fmtc(_itc))
            sc4.metric("Round off",     fmtc(p.get("round_off") or 0))
            sc5.metric("Gross total",   fmtc(p["gross"]))

            # Confirm / Discard
            ac1, ac2 = st.columns(2)
            if ac1.button("✅  Add bill to register", type="primary", use_container_width=True, key=f"confirm_{mk}"):
                if not p.get("supplier") or not p.get("invno") or not p.get("inv_date"):
                    st.error("Supplier, Invoice no., and Date are required.")
                elif (p["val18"]+p["val5"]+p["val12"]+p["exempt"])==0:
                    st.error("Enter at least one taxable/exempt amount.")
                else:
                    try:
                        pts = p["inv_date"].split("-")
                        p["inv_date_obj"] = datetime.date(int(pts[2]),int(pts[1]),int(pts[0]))
                    except: p["inv_date_obj"] = None
                    p["filename"] = st.session_state.get(f"calc_photo_{mk}") or f"manual_{p['invno']}"
                    p["status"] = "ok"; p["source"] = "manual"
                    # Persist new supplier to master if not already there
                    if p.get("supplier") and p.get("gstin"):
                        if p["supplier"] not in st.session_state.suppliers:
                            add_supplier(p["supplier"], p["gstin"])
                    # Clear cached new-supplier text input values
                    for _k in [k for k in st.session_state if k.endswith("_nn_val") or k.endswith("_ng_val")]:
                        del st.session_state[_k]
                    st.session_state.bills[mk].append(p)
                    db_save_bills(mk, st.session_state.bills[mk])
                    del st.session_state[f"calc_{mk}"]
                    st.success(f"✓ Added: {p['supplier']} — {p['invno']}"); st.rerun()
            if ac2.button("✏️  Edit — go back", use_container_width=True, key=f"discard_{mk}"):
                del st.session_state[f"calc_{mk}"]; st.rerun()

        st.markdown("---")
        st.markdown("### 📷 Scan from photos (Gemini AI)")
        if not st.session_state.gemini_key:
            st.warning("Enter Gemini API key in sidebar to use AI scanning.")
        else:
            uploaded=st.file_uploader("Upload bill photos — multiple OK",
                type=["jpg","jpeg","png","webp"],accept_multiple_files=True,key=f"scan{mk}")
            if uploaded:
                existing={b.get("filename") for b in st.session_state.bills[mk]}
                new_files=[f for f in uploaded if f.name not in existing]
                if new_files:
                    prog=st.progress(0,text="Reading bills…")
                    for i,f in enumerate(new_files):
                        prog.progress(i/len(new_files),text=f"Reading {f.name} ({i+1}/{len(new_files)})")
                        try:
                            data=extract_bill_ai(f.read(),f.type,st.session_state.gemini_key)
                            data["filename"]=f.name; data["status"]="ok"; data["source"]="ai"
                        except Exception as e:
                            data={"filename":f.name,"status":"error","error":str(e),"source":"ai"}
                        st.session_state.bills[mk].append(data)
                    db_save_bills(mk, st.session_state.bills[mk])
                    prog.progress(1.0,text=f"✓ Done — {len(new_files)} bill(s). Review below.")
                    st.rerun()

        bills_list=st.session_state.bills.get(mk,[])
        ok_list =[b for b in bills_list if b.get("status")=="ok"]
        err_list=[b for b in bills_list if b.get("status")=="error"]
        for b in err_list:
            st.error(f"❌ {b['filename']}: {b.get('error','unknown')}")

        if ok_list:
            st.markdown("---")
            bs=bills_summary(mk)

            # ── Summary metrics ────────────────────────────────────────────
            c1,c2,c3,c4=st.columns(4)
            c1.metric("Bills entered",    bs["count"])
            c2.metric("Total taxable",    fmtc(bs["p18"]+bs["p5"]+bs["p12"]))
            c3.metric("ITC available",    fmtc(bs["itc"]))
            c4.metric("Gross total",      fmtc(bs["gross"]))

            # ── Quick-view table — visible immediately ─────────────────────
            st.markdown(f"### 📋 Bills added — {len(ok_list)} total")
            tbl_rows = []
            for i, bill in enumerate([b for b in bills_list if b.get("status")=="ok"]):
                itc_b = r2((bill.get("cgst9") or 0)+(bill.get("sgst9") or 0)+
                           (bill.get("cgst25") or 0)+(bill.get("sgst25") or 0))
                tbl_rows.append({
                    "#":            i+1,
                    "Voucher No.":  bill.get("pur_vno","—"),
                    "Supplier":     bill.get("supplier","?"),
                    "GSTIN":        bill.get("gstin","—"),
                    "Supplier Inv.":bill.get("invno","—"),
                    "Date":         bill.get("inv_date","—"),
                    "@18%":         fmtc(bill.get("val18") or 0) if (bill.get("val18") or 0)>0 else "—",
                    "@5%":          fmtc(bill.get("val5")  or 0) if (bill.get("val5")  or 0)>0 else "—",
                    "@12%":         fmtc(bill.get("val12") or 0) if (bill.get("val12") or 0)>0 else "—",
                    "Exempt":       fmtc(bill.get("exempt")or 0) if (bill.get("exempt")or 0)>0 else "—",
                    "Gross":        fmtc(bill.get("gross") or 0),
                    "ITC":          fmtc(itc_b),
                    "Source":       "✏️ Manual" if bill.get("source")=="manual" else "🤖 AI",
                })
            st.dataframe(tbl_rows, use_container_width=True, hide_index=True)

            # ── Edit / remove individual bills ─────────────────────────────
            st.markdown("**Edit or remove a bill:**")
            to_del   = None
            to_save  = None
            for idx, bill in enumerate(st.session_state.bills[mk]):
                if bill.get("status") != "ok": continue
                src_icon = "✏️" if bill.get("source") == "manual" else "🤖"
                lbl2 = (f"{src_icon} #{idx+1} · **{bill.get('supplier','?')}** "
                    f"· {bill.get('invno','?')} · {bill.get('inv_date','?')} "
                    f"· {fmtc(bill.get('gross',0))}")
                with st.expander(lbl2, expanded=False):
                    # bill_fields rendered OUTSIDE form so all widgets update live
                    updated = bill_fields(bill, f"ef{mk}{idx}")
                    # Save / Remove as plain buttons (no form needed)
                    bc1, bc2 = st.columns(2)
                    if bc1.button("💾  Save changes", key=f"sav{mk}{idx}",
                                  type="primary", use_container_width=True):
                        to_save = (idx, updated)
                    if bc2.button("🗑  Remove this bill", key=f"del{mk}{idx}",
                                  use_container_width=True):
                        to_del = idx

            if to_save is not None:
                sidx, supdated = to_save
                supdated["status"] = "ok"
                st.session_state.bills[mk][sidx] = supdated
                db_save_bills(mk, st.session_state.bills[mk])
                st.success(f"✓ Bill #{sidx+1} updated.")
                st.rerun()
            if to_del is not None:
                st.session_state.bills[mk].pop(to_del)
                db_save_bills(mk, st.session_state.bills[mk])
                st.rerun()

            st.markdown("")
            if st.button("🗑 Clear ALL bills for this month", key=f"clr{mk}"):
                st.session_state.bills[mk] = []
                db_save_bills(mk, [])
                st.rerun()
        else:
            st.info("No bills yet. Use the form above or upload photos.")

    # ── TAB 2: DERIVED SALES ──────────────────────────────────────────────────
    with tab_sales:
        bs=bills_summary(mk)
        if bs["count"]==0:
            st.info("Enter purchase bills in the Purchase Bills tab first."); return

        totals=derive_sales_totals(mk)
        st.markdown("Sales are auto-derived from your purchase data.")
        st.caption(f"Ratios used: 18%→18% ×{RATIO_18_TO_18} · 18%→12% ×{RATIO_18_TO_12} · 5%→5% ×{RATIO_5_TO_5} · Exempt ×{RATIO_EX_TO_EX} (from Oct-2025 actuals)")

        c1,c2,c3,c4,c5=st.columns(5)
        c1.metric("Est. total sales",fmtc(sum(totals.values())))
        c2.metric("@18%",fmtc(totals["v18"])); c3.metric("@5%",fmtc(totals["v5"]))
        c4.metric("@12%",fmtc(totals["v12"])); c5.metric("Exempt",fmtc(totals["vex"]))

        st.markdown("---")
        st.markdown("**Daily breakdown**")
        entries=build_daily_sales(mk)
        rows=[{"Date":e["date"].strftime("%d %b"),"Day":e["date"].strftime("%a"),
            "@18%":fmtc(e["v18"]) if e["v18"] else "—",
            "@5%": fmtc(e["v5"])  if e["v5"]  else "—",
            "@12%":fmtc(e["v12"]) if e["v12"] else "—",
            "Exempt":fmtc(e["vex"]) if e["vex"] else "—",
            "Total":fmtc(e["v18"]+e["v5"]+e["v12"]+e["vex"]),
            "Bills":e["nbills"]} for e in entries]
        st.dataframe(rows,use_container_width=True,hide_index=True,height=400)

        st.markdown("---")
        with st.expander("✏️ Override a specific day"):
            days=ld(m,y)
            ov_d=st.number_input("Day", min_value=1, max_value=days, value=1, step=1, key=f"ovd{mk}")
            oc1,oc2,oc3,oc4,oc5=st.columns(5)
            ov18=oc1.number_input("@18%", 0.0,step=50.0,format="%.2f",key=f"ov18{mk}")
            ov5 =oc2.number_input("@5%",  0.0,step=10.0,format="%.2f",key=f"ov5{mk}")
            ov12=oc3.number_input("@12%", 0.0,step=10.0,format="%.2f",key=f"ov12{mk}")
            ovex=oc4.number_input("Exempt",0.0,step=50.0,format="%.2f",key=f"ovex{mk}")
            ovnb=oc5.number_input("Bills", min_value=1, max_value=30, value=4, step=1, key=f"ovnb{mk}")
            if st.button("Save override",key=f"ovsave{mk}"):
                ov_data = {"v18":ov18,"v5":ov5,"v12":ov12,"vex":ovex,"nbills":int(ovnb)}
                st.session_state.sales_override[mk][int(ov_d)] = ov_data
                db_save_override(mk, int(ov_d), ov_data)
                st.success(f"Saved for day {ov_d}"); st.rerun()
            ovs=st.session_state.sales_override.get(mk,{})
            if ovs:
                st.caption(f"{len(ovs)} day(s) overridden: {sorted(ovs.keys())}")
                if st.button("Clear all overrides",key=f"ovclr{mk}"):
                    st.session_state.sales_override[mk]={}
                    db_delete_overrides(mk)
                    st.rerun()

        # ── SMART SALES OVERRIDE ──────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 🎯 Smart Sales Override")
        st.caption(
            "Set a desired gross sales total for this month. The tool distributes it "
            "across all 31 days using ±30 % daily variation with SLVT's real sales pattern: "
            "75 % of gross → 18 % GST slab, 25 % of gross → 5 % GST slab, zero Exempt "
            "(exempt items are one-off entries, not daily). "
            "This overwrites all existing day-overrides and is reflected in both the "
            "Sales Register (.xlsx) and GSTR-1 JSON."
        )

        _saved_target = db_load_target(mk)
        _default_target = float(_saved_target) if _saved_target else 1600000.0

        _target_col, _btn_col = st.columns([3, 1])
        smart_target = _target_col.number_input(
            "Desired Total Sales (₹)",
            min_value=10000.0,
            max_value=50000000.0,
            value=_default_target,
            step=10000.0,
            format="%.2f",
            key=f"smart_target_{mk}",
            help="Enter the gross sales amount you want this month to show (e.g. 1600000 for ₹16 Lakhs)."
        )

        # Show breakdown preview  (75% @18%, 25% @5%, no exempt on regular days)
        _rev18 = smart_target * 0.75
        _rev5  = smart_target - _rev18
        _tx18  = _rev18 / 1.18
        _tx5   = _rev5  / 1.05
        st.markdown(
            f"<div style='background:#eaf7ea;border-radius:6px;padding:8px 14px;"
            f"font-size:12px;color:#1a3c1a;margin:4px 0 8px 0'>"
            f"Split preview → "
            f"<b>18% slab (75%):</b> taxable {fmtc(_tx18)} + GST {fmtc(_tx18*0.18)} = {fmtc(_rev18)} &nbsp;|&nbsp; "
            f"<b>5% slab (25%):</b> taxable {fmtc(_tx5)} + GST {fmtc(_tx5*0.05)} = {fmtc(_rev5)} &nbsp;|&nbsp; "
            f"<b>Grand total gross: {fmtc(smart_target)}</b> &nbsp;|&nbsp; "
            f"<b>Exempt: ₹0</b> (exempt items entered separately if any)"
            f"</div>",
            unsafe_allow_html=True
        )

        if _btn_col.button(
            "🚀 Generate Smart Override",
            key=f"smart_gen_{mk}",
            type="primary",
            use_container_width=True
        ):
            with st.spinner(f"Distributing ₹{smart_target:,.0f} across {ld(m,y)} days…"):
                distribute_target_sales(mk, smart_target)
            st.success(
                f"✅ Smart override applied! All {ld(m,y)} days now total "
                f"**{fmtc(smart_target)}**. "
                f"Download the Sales Register or GSTR-1 JSON from the Documents tab."
            )
            st.rerun()

        if _saved_target:
            st.caption(f"Last saved target for this month: **{fmtc(_saved_target)}**")

    # ── TAB 3: DOWNLOAD DOCUMENTS ─────────────────────────────────────────────
    with tab_docs:
        bs=bills_summary(mk)
        if bs["count"]==0:
            st.info("Enter purchase bills first."); return

        start_vno=get_voucher_start(mk)
        entries  =build_daily_sales(mk)
        purchases=ok_bills(mk)
        sales_xls,sales_g,last_vno=make_sales_xlsx(mk,entries,start_vno)
        pur_xls,_ =make_purchase_xlsx(mk,purchases)
        gstr1_dict=make_gstr1_json(mk,sales_g,start_vno,last_vno,last_vno-start_vno+1)
        gstr1_json=json.dumps(gstr1_dict,indent=2)
        m2,y2=parse_mk(mk); mname=lbl.replace(" ","_")

        st.markdown(f"### Summary — {lbl}")
        out_gst=r2(sales_g["c9"]+sales_g["s9"]+sales_g["c25"]+sales_g["s25"]+sales_g["c6"]+sales_g["s6"])
        net_pay=r2(max(0, out_gst - bs["itc"]))
        c1,c2,c3,c4=st.columns(4)
        c1.metric("Total sales",     fmtc(sales_g["val"]))
        c2.metric("Output GST",      fmtc(out_gst))
        c3.metric("ITC (purchases)", fmtc(bs["itc"]))
        c4.metric("Net GST payable", fmtc(net_pay))

        st.markdown("---")
        st.markdown("### Download all 3 documents")

        st.markdown("**1. Sales Register** — Tally-format daily cash sales")
        st.download_button(f"⬇  Sales_Register_{mname}.xlsx",data=sales_xls,
            file_name=f"Sales_Register_{mname}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",use_container_width=True,key=f"dl_s_{mk}")

        st.markdown("**2. Purchase Register** — Tally-format supplier bills")
        st.download_button(f"⬇  Purchase_Register_{mname}.xlsx",data=pur_xls,
            file_name=f"Purchase_Register_{mname}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",use_container_width=True,key=f"dl_p_{mk}")

        st.markdown("**3. GSTR-1 JSON** — upload at gstn.gov.in → Returns → GSTR-1 → Upload JSON")
        st.download_button(f"⬇  {GSTIN}_GSTR1_{MON_ABBR[m2]}{y2}.json",
            data=gstr1_json.encode(),
            file_name=f"{GSTIN}_GSTR1_{MON_ABBR[m2]}{y2}.json",
            mime="application/json",
            type="primary",use_container_width=True,key=f"dl_g_{mk}")

        st.info(f"Vouchers: **{start_vno}** → **{last_vno}** ({last_vno-start_vno+1} bills) · Next month starts at **{last_vno+1}**")

        with st.expander("Preview GSTR-1 JSON"):
            t1,t2,t3,t4=st.tabs(["B2CS","Nil/Exempt","HSN Summary","Doc Series"])
            with t1:
                for row in gstr1_dict.get("b2cs",[]):
                    st.markdown(f"**{row['rt']}%** — Taxable: {fmtc(row['txval'])} · CGST: {fmtc(row['camt'])} · SGST: {fmtc(row['samt'])}")
            with t2:
                nil=gstr1_dict["nil"]["inv"][0]
                st.markdown(f"Exempt: {fmtc(nil['expt_amt'])} · Nil: {fmtc(nil['nil_amt'])}")
        with t3:
            hsn_rows = st.session_state.get("hsn_entries", {}).get(mk, [])
            is_filed = mk in FILED_HSN and bool(FILED_HSN[mk])

            if is_filed:
                st.success("✅ HSN data loaded from your filed GSTR-1 JSON — will be included as-is in the download.")
            elif mk == "03-2026":
                st.info("📝 March HSN — enter your HSN entries below. They'll be saved and included in the JSON.")
            else:
                st.warning("⚠ No HSN data. JSON will be downloaded without HSN section.")

            # ── Validation: compare HSN totals to B2CS ──────────────────────
            if hsn_rows:
                from collections import defaultdict
                hsn_by_rate = defaultdict(float)
                for row in hsn_rows:
                    hsn_by_rate[row["rt"]] += row["txval"]
                b2cs_by_rate = {row["rt"]: row["txval"] for row in gstr1_dict.get("b2cs",[])}
                b2cs_ex = gstr1_dict["nil"]["inv"][0]["expt_amt"]

                st.markdown("**HSN vs B2CS Validation**")
                # Tolerance: 0.15% of B2CS value (or ₹5 minimum).
                # Your filed Nov–Feb JSONs have HSN 18% overshooting B2CS by up to 0.1%
                # (e.g. Jan: +₹398.70 on ₹3,98,700) — the portal accepted all of them.
                # A diff within 0.15% is normal rounding across HSN lines and is portal-safe.
                all_ok = True
                has_warned = False
                for rt in sorted(set(list(hsn_by_rate.keys()) + list(b2cs_by_rate.keys()))):
                    if rt == 0:
                        hsn_v = round(hsn_by_rate.get(0, 0.0), 2)
                        b_v   = b2cs_ex
                    else:
                        hsn_v = round(hsn_by_rate.get(rt, 0.0), 2)
                        b_v   = b2cs_by_rate.get(rt, 0.0)
                    diff      = abs(hsn_v - b_v)
                    tolerance = max(5.0, b_v * 0.0015)   # 0.15% or ₹5, whichever is larger
                    ok        = diff <= tolerance
                    if not ok: all_ok = False
                    label = "Exempt" if rt == 0 else f"{rt}%"
                    if diff == 0:
                        st.markdown(f"✅ **{label}** — HSN: {fmtc(hsn_v)} | B2CS: {fmtc(b_v)}")
                    elif ok:
                        st.markdown(f"✅ **{label}** — HSN: {fmtc(hsn_v)} | B2CS: {fmtc(b_v)} | Diff: {fmtc(diff)} *(within portal tolerance)*")
                    else:
                        st.markdown(f"❌ **{label}** — HSN: {fmtc(hsn_v)} | B2CS: {fmtc(b_v)} | Diff: {fmtc(diff)} *(too large — fix needed)*")
                if all_ok:
                    st.success("✅ JSON is portal-ready. All HSN vs B2CS differences are within the accepted tolerance.")
                else:
                    st.error("❌ HSN total differs from B2CS by more than 0.15%. This may cause portal rejection — edit HSN rows below.")

                # ── Show HSN table ───────────────────────────────────────────
                st.markdown("**HSN entries:**")
                import pandas as pd
                df_rows = [{"#":r["num"],"HSN":r["hsn_sc"],"UQC":r["uqc"],"Qty":r["qty"],
                            "Rate%":r["rt"],"Taxable ₹":r["txval"],
                            "CGST ₹":r["camt"],"SGST ₹":r["samt"]} for r in hsn_rows]
                st.dataframe(df_rows, use_container_width=True, hide_index=True)

            # ── HSN editor (only shown for March or empty months) ───────────
            if not is_filed:
                st.markdown("---")
                st.markdown("**Add/Edit HSN entries for March:**")
                cur_rows = list(st.session_state.hsn_entries.get(mk, []))
                # Remove button for existing rows
                to_del_hsn = None
                for i, row in enumerate(cur_rows):
                    rc1,rc2,rc3,rc4,rc5,rc6,rc7 = st.columns([1,1.5,1,1,1.5,1.5,0.7])
                    row["hsn_sc"] = rc1.text_input("HSN",  value=row.get("hsn_sc",""), key=f"hsnc_{mk}_{i}")
                    row["uqc"]    = rc2.selectbox("UQC", ["PCS","PAC","KGS","LTR","NOS","BAG"], 
                                                  index=["PCS","PAC","KGS","LTR","NOS","BAG"].index(row.get("uqc","PCS")) 
                                                  if row.get("uqc","PCS") in ["PCS","PAC","KGS","LTR","NOS","BAG"] else 0,
                                                  key=f"hsnuqc_{mk}_{i}")
                    row["qty"]    = rc3.number_input("Qty",    value=int(row.get("qty",1)), min_value=0, step=1, key=f"hsnqty_{mk}_{i}")
                    row["rt"]     = rc4.number_input("Rate%",  value=int(row.get("rt",18)), min_value=0, max_value=28, step=1, key=f"hsnrt_{mk}_{i}")
                    row["txval"]  = rc5.number_input("Taxable",value=float(row.get("txval",0)), min_value=0.0, step=0.01, format="%.2f", key=f"hsntx_{mk}_{i}")
                    row["camt"]   = r2(row["txval"] * row["rt"] / 100 / 2)
                    row["samt"]   = row["camt"]
                    rc6.metric("CGST/SGST", fmtc(row["camt"]))
                    if rc7.button("🗑", key=f"hsndel_{mk}_{i}"):
                        to_del_hsn = i
                    cur_rows[i] = row

                if to_del_hsn is not None:
                    cur_rows.pop(to_del_hsn)
                    st.session_state.hsn_entries[mk] = cur_rows
                    db_save_hsn(mk, cur_rows)
                    st.rerun()

                # Add new row
                st.markdown("**Add a new HSN row:**")
                na1,na2,na3,na4,na5 = st.columns([1.5,1,1,1.5,0.8])
                new_hsn = na1.text_input("HSN Code", key=f"nhsn_{mk}", placeholder="e.g. 38089290")
                new_uqc = na2.selectbox("UQC", ["PCS","PAC","KGS","LTR","NOS","BAG"], key=f"nuqc_{mk}")
                new_qty = na3.number_input("Qty", min_value=0, step=1, key=f"nqty_{mk}")
                new_rt  = na4.number_input("Rate%", min_value=0, max_value=28, step=1, value=18, key=f"nrt_{mk}")
                new_tx  = na5.number_input("Taxable ₹", min_value=0.0, step=0.01, format="%.2f", key=f"ntx_{mk}")
                if st.button("➕ Add HSN row", key=f"addhsn_{mk}"):
                    if new_hsn and new_tx > 0:
                        new_camt = r2(new_tx * new_rt / 100 / 2)
                        new_row  = {"num": len(cur_rows)+1, "hsn_sc": new_hsn,
                                    "desc":"","user_desc":"",
                                    "uqc": new_uqc, "qty": new_qty, "rt": new_rt,
                                    "txval": new_tx, "iamt":0,
                                    "camt": new_camt, "samt": new_camt, "csamt":0}
                        cur_rows.append(new_row)
                        st.session_state.hsn_entries[mk] = cur_rows
                        db_save_hsn(mk, cur_rows)
                        st.success("HSN row added."); st.rerun()
                    else:
                        st.error("Enter HSN code and taxable value.")

                if cur_rows and st.button("💾 Save HSN changes", key=f"savhsn_{mk}", type="primary"):
                    st.session_state.hsn_entries[mk] = cur_rows
                    db_save_hsn(mk, cur_rows)
                    st.success("✅ HSN entries saved."); st.rerun()
            with t4:
                doc=gstr1_dict["doc_issue"]["doc_det"][0]["docs"][0]
                st.markdown(f"From **{doc['from']}** to **{doc['to']}** · Issued: **{doc['net_issue']}**")

# ─────────────────────────────────────────────────────────────────────────────

def gst_ss_init():
    if "gst_bills" not in st.session_state:
        st.session_state.gst_bills = {}
        for mk, _ in MONTHS:
            st.session_state.gst_bills[mk] = db_load_bills(mk)
    if "gst_overrides" not in st.session_state:
        st.session_state.gst_overrides = {}
        for mk, _ in MONTHS:
            st.session_state.gst_overrides[mk] = db_load_overrides(mk)
    if "gst_suppliers" not in st.session_state:
        merged = dict(DEFAULT_SUPPLIERS)
        merged.update(db_load_suppliers())
        st.session_state.gst_suppliers = merged
    if "gemini_key" not in st.session_state:
        st.session_state.gemini_key = ""
    if "gst_sub_page" not in st.session_state:
        st.session_state.gst_sub_page = "dashboard"
    # Sync HSN entries (seeded by ss_init)
    if "hsn_entries" not in st.session_state:
        st.session_state.hsn_entries = {}
        for mk, _ in MONTHS:
            from_db = db_load_hsn(mk)
            if from_db is not None:
                st.session_state.hsn_entries[mk] = from_db
            elif mk in FILED_HSN and FILED_HSN[mk]:
                st.session_state.hsn_entries[mk] = list(FILED_HSN[mk])
                db_save_hsn(mk, FILED_HSN[mk])
            else:
                st.session_state.hsn_entries[mk] = []

if page == "🏠 Dashboard":
    st.subheader("📊 Dashboard Overview")
    c1, c2, c3, c4 = st.columns(4)

    total_prods    = conn.execute("SELECT COUNT(*) FROM inventory").fetchone()[0]
    stock_val      = conn.execute("SELECT COALESCE(SUM(quantity*mrp),0) FROM inventory").fetchone()[0]
    low_count      = conn.execute(
        "SELECT COUNT(*) FROM inventory WHERE quantity<?", (LOW_STOCK_ALERT,)).fetchone()[0]
    today_sales    = conn.execute(
        "SELECT COALESCE(SUM(total_amount),0) FROM invoices WHERE DATE(invoice_date)=?",
        (datetime.date.today().isoformat(),)).fetchone()[0]
    total_barcodes = conn.execute("SELECT COUNT(*) FROM qr_registry").fetchone()[0]

    c1.metric("🛒 Products",      total_prods)
    c2.metric("💰 Stock Value",   f"₹ {stock_val:,.0f}")
    c3.metric("⚠️ Low Stock",     low_count)
    c4.metric("📈 Today's Sales", f"₹ {today_sales:,.2f}")

    st.info(f"🔬 **{total_barcodes}** QR codes registered in system")

    if low_count > 0:
        st.markdown("---")
        st.warning(f"⚠️ **{low_count} item(s) below {LOW_STOCK_ALERT} units — reorder needed!**")
        for item in conn.execute(
            "SELECT name,section,row_no,slot,quantity,unit FROM inventory WHERE quantity<? ORDER BY quantity",
            (LOW_STOCK_ALERT,)
        ).fetchall():
            st.markdown(
                f"🔴 **{item['name']}** | "
                f"{item['section']} › {item['row_no']} › Slot {item['slot']} | "
                f"Stock: **{item['quantity']} {item['unit']}**"
            )

    st.markdown("---")
    st.subheader("🧾 Recent Invoices")
    import pandas as pd
    recent = conn.execute(
        "SELECT invoice_no,invoice_date,customer_name,total_amount FROM invoices ORDER BY id DESC LIMIT 10"
    ).fetchall()
    if recent:
        df = pd.DataFrame([dict(r) for r in recent])
        df.columns = ["Invoice No","Date","Customer","Total (₹)"]
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.info("No invoices yet. Create your first bill!")

# ═════════════════════════════════════════════════════════════════════════════
# PAGE — INVENTORY
# ═════════════════════════════════════════════════════════════════════════════
elif page == "📦 Inventory":
    st.subheader("📦 Inventory Management")
    tab1, tab2 = st.tabs(["📋 View Inventory", "➕ Add / Edit Product"])

    with tab1:
        import pandas as pd
        products = conn.execute(
            "SELECT * FROM inventory ORDER BY section,row_no,slot"
        ).fetchall()
        if products:
            rows = []
            for p in products:
                rows.append({
                    "ID":       p["id"],
                    "QR Code":  p["qr_code"] or "—",
                    "Name":     p["name"],
                    "HSN":      p["hsn_code"],
                    "GST":      f"{(p['gst_rate'] or 0)*100:.0f}%",
                    "Section":  p["section"],
                    "Row":      p["row_no"],
                    "Slot":     p["slot"],
                    "Qty":      p["quantity"],
                    "Unit":     p["unit"],
                    "MRP(₹)":   p["mrp"],
                    "Status":   "🔴 LOW" if p["quantity"] < LOW_STOCK_ALERT else "🟢 OK",
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        else:
            st.info("No products yet.")

    with tab2:
        st.markdown("#### Add New Product Manually")
        with st.form("add_product_form"):
            c1, c2 = st.columns(2)
            p_barcode = c1.text_input("QR Code (scan or type)",
                                      placeholder="Leave blank if no QR code")
            p_name    = c2.text_input("Product Name *")
            p_hsn     = c1.text_input("HSN Code *")
            p_section = c2.selectbox("Section", [
                "Chemical Section","Organic Section","Micro-Nutrient",
                "Mineral Section","Pesticide Section","Other"])
            p_row  = c1.text_input("Row No (e.g. Row 1)")
            p_slot = c2.text_input("Slot (e.g. AA)")
            p_unit = c1.selectbox("Unit", ["Kg","Bag","Litre","Gram","Nos"])
            p_qty  = c2.number_input("Opening Stock Qty", min_value=0.0, step=0.5)
            p_mrp  = c1.number_input("MRP per Unit (₹)", min_value=0.0, step=0.5)
            p_cost = c2.number_input("Cost Price per Unit (₹)", min_value=0.0, step=0.5)
            p_gst_label = c1.selectbox("GST Rate *", list(GST_RATES.keys()),
                          index=list(GST_RATES.keys()).index(DEFAULT_GST_LABEL))

            if st.form_submit_button("➕ Add Product", type="primary"):
                if p_name and p_hsn and p_row and p_slot:
                    try:
                        conn.execute("""
                            INSERT INTO inventory
                                (qr_code,name,hsn_code,section,row_no,slot,
                                 unit,quantity,mrp,cost_price,gst_rate)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?)
                        """, (p_barcode.strip() or None, p_name, p_hsn, p_section,
                              p_row, p_slot, p_unit, p_qty, p_mrp, p_cost,
                              GST_RATES[p_gst_label]))
                        conn.commit()
                        if p_barcode.strip():
                            pid = conn.execute(
                                "SELECT id FROM inventory WHERE qr_code=?",
                                (p_barcode.strip(),)
                            ).fetchone()["id"]
                            register_barcode_scan(p_barcode.strip(), pid, p_name)
                        st.success(f"✅ '{p_name}' added successfully!")
                        st.rerun()
                    except sqlite3.IntegrityError:
                        st.error("❌ QR code already registered for another product!")
                else:
                    st.error("Fill all required (*) fields.")

        st.markdown("---")
        st.markdown("#### 🔄 Restock a Product")
        all_prods = conn.execute(
            "SELECT id,name,quantity,unit FROM inventory ORDER BY name"
        ).fetchall()
        prod_opts = {f"{p['name']} (Stock: {p['quantity']} {p['unit']})": p["id"]
                     for p in all_prods}
        if prod_opts:
            sel     = st.selectbox("Select Product", list(prod_opts.keys()))
            add_qty = st.number_input("Add Quantity", min_value=0.0, step=0.5)
            if st.button("✅ Update Stock"):
                conn.execute("UPDATE inventory SET quantity=quantity+? WHERE id=?",
                             (add_qty, prod_opts[sel]))
                conn.commit()
                st.success("Stock updated!")
                st.rerun()

# ═════════════════════════════════════════════════════════════════════════════
# PAGE — BARCODE SCANNER
# ═════════════════════════════════════════════════════════════════════════════
elif page == "🔬 QR Scanner":
    st.subheader("🔬 USB QR Code Scanner — Inventory & Billing")

    # ── Instructions banner ────────────────────────────────────────────────
    st.markdown("""
    <div style="background:#e3f2fd;border-left:5px solid #1565c0;
                padding:12px 16px;border-radius:8px;margin-bottom:16px;">
        <b>📡 How to use your USB QR Code Scanner:</b><br>
        1. Plug the USB QR scanner into your PC — it works like a keyboard, no drivers needed<br>
        2. Select the <b>mode</b> below (Receive Stock or Billing)<br>
        3. Click inside the <b>"Scan QR Code Here"</b> box<br>
        4. Point the scanner at any product QR code and pull the trigger<br>
        5. The scanner reads the QR and types the code automatically<br><br>
        <b>✅ Known product</b> → Shows details + restock/billing form instantly<br>
        <b>🆕 New product</b> → Prompts you to fill product details and register it
    </div>
    """, unsafe_allow_html=True)

    # ── Mode selector ──────────────────────────────────────────────────────
    scan_mode = st.radio(
        "Scanner Mode",
        ["📥 Receive Stock (Add to Inventory)", "🧾 Billing (Add to Cart)"],
        horizontal=True
    )
    st.markdown("---")

    # ── SCAN INPUT — USB scanner types into this box & presses Enter ───────
    with st.form("qr_form", clear_on_submit=True):
        scanned   = st.text_input(
            "🔍 Scan QR Code Here",
            placeholder="Click here, then scan product QR code with USB scanner...",
            help="The USB QR scanner acts as a keyboard — it reads and types the QR code automatically"
        )
        col_scan, col_manual = st.columns([1, 3])
        submitted = col_scan.form_submit_button("🔎 Lookup", type="primary")
        col_manual.caption("💡 Tip: Scanner auto-submits on Enter. Manual lookup button is a backup.")

    # Process scan
    if submitted and scanned.strip():
        barcode = scanned.strip()
        product, is_new = lookup_barcode(barcode)
        st.session_state.qr_result   = product
        st.session_state.qr_is_new   = is_new
        st.session_state.last_qr_code = barcode

    # ── RESULT DISPLAY ─────────────────────────────────────────────────────
    barcode = st.session_state.last_qr_code
    product = st.session_state.qr_result
    is_new  = st.session_state.qr_is_new

    if barcode:

        # ══════════════════════════════════════════════════════════════════
        # CASE A: KNOWN PRODUCT
        # ══════════════════════════════════════════════════════════════════
        if not is_new and product:
            gst_pct     = (product.get("gst_rate") or 0.05) * 100
            stock_color = "🔴" if product["quantity"] < LOW_STOCK_ALERT else "🟢"

            # Rich product card
            st.markdown(f"""
            <div class="scan-found">
                <h3>✅ Product Found in Inventory!</h3>
                <div class="detail-row">
                    <span class="detail-pill">📦 <b>{product['name']}</b></span>
                    <span class="detail-pill">📱 QR Code: <b>{barcode}</b></span>
                    <span class="detail-pill">HSN: {product['hsn_code']}</span>
                    <span class="detail-pill">GST: {gst_pct:.0f}%</span>
                </div>
                <div class="detail-row">
                    <span class="detail-pill">📍 {product['section']} › {product['row_no']} › Slot <b>{product['slot']}</b></span>
                    <span class="detail-pill">{stock_color} Stock: <b>{product['quantity']} {product['unit']}</b></span>
                    <span class="detail-pill">💰 MRP: ₹{product['mrp']:.2f}/{product['unit']}</span>
                    <span class="detail-pill">🏭 Cost: ₹{product['cost_price']:.2f}</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

            if product["quantity"] < LOW_STOCK_ALERT:
                st.warning(f"⚠️ Low Stock Alert! Only **{product['quantity']} {product['unit']}** remaining.")

            register_barcode_scan(barcode, product["id"], product["name"])

            # ── RECEIVE STOCK MODE ─────────────────────────────────────────
            if "Receive Stock" in scan_mode:
                st.markdown("#### 📥 Update Stock for this Product")
                with st.form("restock_form"):
                    rc1, rc2, rc3 = st.columns(3)
                    r_qty  = rc1.number_input("Quantity Received",   min_value=0.1, step=0.5, value=1.0)
                    r_cost = rc2.number_input("Updated Cost (₹/unit)", min_value=0.0, step=0.5,
                                              value=float(product["cost_price"]))
                    r_mrp  = rc3.number_input("Updated MRP (₹/unit)",  min_value=0.0, step=0.5,
                                              value=float(product["mrp"]))
                    if st.form_submit_button("✅ Confirm Stock Receipt", type="primary"):
                        conn.execute("""
                            UPDATE inventory
                            SET quantity=quantity+?, cost_price=?, mrp=?
                            WHERE id=?
                        """, (r_qty, r_cost, r_mrp, product["id"]))
                        conn.commit()
                        new_qty = product["quantity"] + r_qty
                        st.success(
                            f"✅ Stock updated! **{product['name']}** → "
                            f"**{new_qty} {product['unit']}** now in stock."
                        )
                        st.session_state.last_qr_code   = ""
                        st.session_state.qr_result      = None
                        st.rerun()

            # ── BILLING MODE ───────────────────────────────────────────────
            else:
                st.markdown("#### 🧾 Add this Product to Bill")
                with st.form("scan_bill_form"):
                    b_qty = st.number_input(
                        f"Quantity to Bill (Available: {product['quantity']} {product['unit']})",
                        min_value=0.1, step=0.5, value=1.0
                    )
                    # Preview line total before adding
                    gst_rate = product.get("gst_rate") or 0.05
                    taxable  = round(product["mrp"] * b_qty, 2)
                    cgst_prev, sgst_prev, total_prev = calculate_gst(taxable, gst_rate)
                    pc1,pc2,pc3,pc4 = st.columns(4)
                    pc1.metric("Taxable",  f"₹ {taxable:.2f}")
                    pc2.metric(f"CGST ({gst_rate/2*100:.1f}%)", f"₹ {cgst_prev:.2f}")
                    pc3.metric(f"SGST ({gst_rate/2*100:.1f}%)", f"₹ {sgst_prev:.2f}")
                    pc4.metric("Line Total", f"₹ {total_prev:.2f}")

                    if st.form_submit_button("➕ Add to Cart", type="primary"):
                        if b_qty > product["quantity"]:
                            st.error(f"❌ Only {product['quantity']} {product['unit']} in stock!")
                        else:
                            cgst, sgst, total = calculate_gst(
                                round(product["mrp"] * b_qty, 2), gst_rate)
                            existing = next((i for i, c in enumerate(st.session_state.cart)
                                             if c["product_id"] == product["id"]), None)
                            if existing is not None:
                                st.session_state.cart[existing]["quantity"] += b_qty
                                c_ = st.session_state.cart[existing]
                                new_tax = round(c_["unit_price"] * c_["quantity"], 2)
                                c_["taxable_value"], c_["cgst_amount"], \
                                c_["sgst_amount"],   c_["line_total"] = \
                                    new_tax, *calculate_gst(new_tax, gst_rate)
                            else:
                                st.session_state.cart.append({
                                    "product_id":    product["id"],
                                    "product_name":  product["name"],
                                    "hsn_code":      product["hsn_code"],
                                    "gst_rate":      gst_rate,
                                    "quantity":      b_qty,
                                    "unit":          product["unit"],
                                    "unit_price":    product["mrp"],
                                    "taxable_value": round(product["mrp"] * b_qty, 2),
                                    "cgst_amount":   cgst,
                                    "sgst_amount":   sgst,
                                    "line_total":    total,
                                })
                            st.success(
                                f"✅ **{product['name']}** × {b_qty} added to cart! "
                                f"Go to 🧾 New Bill to complete the invoice."
                            )
                            st.session_state.last_qr_code   = ""
                            st.session_state.qr_result      = None
                            st.rerun()

        # ══════════════════════════════════════════════════════════════════
        # CASE B: BRAND NEW PRODUCT — first time this QR code is scanned
        # ══════════════════════════════════════════════════════════════════
        elif is_new:
            st.markdown(f"""
            <div class="scan-new">
                <h3>🆕 New Product Detected — First Time Scan!</h3>
                <p>
                    QR Code <b>{barcode}</b> is <b>not yet registered</b> in your inventory.<br>
                    This is the <b>first time</b> this product is being added to Sri Lakshmi Venkateshwara Traders.<br>
                    Fill in the details below and click <b>Register Product</b> to save it permanently.
                </p>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("#### 📋 Register New Product Details")
            st.info(f"📱 QR Code **`{barcode}`** will be linked to this product automatically.")

            with st.form("new_product_scan_form"):
                fc1, fc2 = st.columns(2)
                np_name    = fc1.text_input("Product Name *",
                                            placeholder="e.g. Urea (46% N)")
                np_hsn     = fc2.text_input("HSN Code *",
                                            placeholder="e.g. 31021010",
                                            help="Find HSN on your purchase bill or GST portal")
                np_section = fc1.selectbox("Section", [
                    "Chemical Section","Organic Section","Micro-Nutrient",
                    "Mineral Section","Pesticide Section","Other"])
                np_row     = fc2.text_input("Row No", placeholder="e.g. Row 1")
                np_slot    = fc1.text_input("Slot",   placeholder="e.g. AA")
                np_unit    = fc2.selectbox("Unit", ["Kg","Bag","Litre","Gram","Nos"])
                np_qty     = fc1.number_input("Opening Stock Qty",    min_value=0.0, step=0.5)
                np_mrp     = fc2.number_input("MRP per Unit (₹)",     min_value=0.0, step=0.5)
                np_cost    = fc1.number_input("Cost Price per Unit (₹)", min_value=0.0, step=0.5)
                np_gst     = fc2.selectbox("GST Rate *", list(GST_RATES.keys()),
                                           index=list(GST_RATES.keys()).index(DEFAULT_GST_LABEL),
                                           help="5% fertilizers | 12% pesticides | 18% micronutrients | 0% organic/seeds")

                # Live GST preview
                st.markdown("---")
                st.markdown("**📊 GST Preview for this Product**")
                if np_mrp > 0 and np_qty > 0:
                    rate      = GST_RATES[np_gst]
                    taxable   = round(np_mrp * np_qty, 2)
                    cgst_p, sgst_p, total_p = calculate_gst(taxable, rate)
                    pv1,pv2,pv3,pv4,pv5 = st.columns(5)
                    pv1.metric("MRP × Qty",       f"₹ {taxable:.2f}")
                    pv2.metric(f"CGST ({rate/2*100:.1f}%)", f"₹ {cgst_p:.2f}")
                    pv3.metric(f"SGST ({rate/2*100:.1f}%)", f"₹ {sgst_p:.2f}")
                    pv4.metric("Total GST",        f"₹ {cgst_p+sgst_p:.2f}")
                    pv5.metric("Invoice Value",    f"₹ {total_p:.2f}")
                else:
                    st.caption("Enter MRP and Opening Qty above to see GST preview.")

                st.markdown("---")
                if st.form_submit_button("✅ Register Product & Save to Inventory", type="primary"):
                    if np_name and np_hsn and np_row and np_slot:
                        try:
                            conn.execute("""
                                INSERT INTO inventory
                                    (qr_code,name,hsn_code,section,row_no,slot,
                                     unit,quantity,mrp,cost_price,gst_rate)
                                VALUES (?,?,?,?,?,?,?,?,?,?,?)
                            """, (barcode, np_name, np_hsn, np_section,
                                  np_row, np_slot, np_unit, np_qty,
                                  np_mrp, np_cost, GST_RATES[np_gst]))
                            conn.commit()

                            new_id = conn.execute(
                                "SELECT id FROM inventory WHERE qr_code=?", (barcode,)
                            ).fetchone()["id"]
                            register_barcode_scan(barcode, new_id, np_name)

                            # Show full summary of what was registered
                            rate    = GST_RATES[np_gst]
                            taxable = round(np_mrp * max(np_qty, 1), 2)
                            cgst_r, sgst_r, total_r = calculate_gst(taxable, rate)

                            st.success("✅ Product successfully registered in inventory!")
                            st.markdown(f"""
                            | Field | Value |
                            |---|---|
                            | **Product Name** | {np_name} |
                            | **QR Code** | `{barcode}` |
                            | **HSN Code** | {np_hsn} |
                            | **Location** | {np_section} › {np_row} › Slot {np_slot} |
                            | **Opening Stock** | {np_qty} {np_unit} |
                            | **MRP** | ₹ {np_mrp:.2f} / {np_unit} |
                            | **Cost Price** | ₹ {np_cost:.2f} / {np_unit} |
                            | **GST Rate** | {rate*100:.0f}% (CGST {rate/2*100:.1f}% + SGST {rate/2*100:.1f}%) |
                            | **Stock Value (MRP)** | ₹ {np_mrp * np_qty:,.2f} |
                            """)

                            st.session_state.last_qr_code   = ""
                            st.session_state.qr_result      = None
                            st.session_state.qr_is_new      = False
                            st.rerun()
                        except sqlite3.IntegrityError:
                            st.error("❌ This QR code is already registered!")
                    else:
                        st.error("Please fill all required (*) fields.")

    # ── Scan History ───────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📜 Scan History (Last 20)")
    import pandas as pd
    history = conn.execute("""
        SELECT br.qr_code, br.product_name, br.first_scanned, br.scan_count,
               i.section, i.row_no, i.slot, i.quantity, i.unit
        FROM qr_registry br
        LEFT JOIN inventory i ON br.product_id = i.id
        ORDER BY br.first_scanned DESC LIMIT 20
    """).fetchall()
    if history:
        df = pd.DataFrame([dict(r) for r in history])
        df.columns = ["QR Code","Product","First Scanned","Times Scanned",
                      "Section","Row","Slot","Current Qty","Unit"]
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.info("No scan history yet. Start scanning products!")

# ═════════════════════════════════════════════════════════════════════════════
# PAGE — NEW BILL  (simplified for ease of use)
# ═════════════════════════════════════════════════════════════════════════════
elif page == "🧾 New Bill":

    import pandas as pd

    # ── Session state for customer fields ─────────────────────────────────
    for k, v in {"bill_cust_name":"","bill_cust_phone":"","bill_cust_gstin":""}.items():
        if k not in st.session_state:
            st.session_state[k] = v

    # ══════════════════════════════════════════════════════════════════════
    # STEP 1 — SCAN BOX  (always visible at top, always focused)
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("""
    <div style="background:#1a3c5e;border-radius:12px;padding:18px 22px;margin-bottom:18px;">
        <h2 style="color:white;margin:0 0 4px 0;font-size:1.4rem;">📱 Step 1 — Scan Product</h2>
        <p style="color:#c8e6c9;margin:0;font-size:0.9rem;">
            Click the box below → point scanner at QR code → item added instantly
        </p>
    </div>
    """, unsafe_allow_html=True)

    with st.form("bill_qr_form", clear_on_submit=True):
        bill_qr = st.text_input(
            "Scan QR Code",
            placeholder="👆 Click here first, then scan the product QR code...",
            label_visibility="collapsed"
        )
        scan_submitted = st.form_submit_button(
            "➕  ADD TO CART",
            type="primary",
            use_container_width=True
        )

    # ── Process scan ───────────────────────────────────────────────────────
    if scan_submitted and bill_qr.strip():
        bc = bill_qr.strip()
        product, is_new = lookup_barcode(bc)

        if is_new or product is None:
            st.error(f"❌  QR code not recognised. Register this product in 🔬 QR Scanner first.")
        elif product["quantity"] <= 0:
            st.error(f"❌  {product['name']} is OUT OF STOCK.")
        else:
            gst_rate = product.get("gst_rate") or 0.05
            existing = next((i for i,c in enumerate(st.session_state.cart)
                             if c["product_id"] == product["id"]), None)
            if existing is not None:
                new_qty = st.session_state.cart[existing]["quantity"] + 1
                if new_qty > product["quantity"]:
                    st.error(f"❌  Only {product['quantity']} {product['unit']} in stock!")
                else:
                    st.session_state.cart[existing]["quantity"] = new_qty
                    new_tax = round(product["mrp"] * new_qty, 2)
                    (st.session_state.cart[existing]["taxable_value"],
                     st.session_state.cart[existing]["cgst_amount"],
                     st.session_state.cart[existing]["sgst_amount"],
                     st.session_state.cart[existing]["line_total"]) = (
                        new_tax, *calculate_gst(new_tax, gst_rate))
                    st.toast(f"✅  {product['name']}  ×{new_qty}", icon="🛒")
            else:
                taxable = round(product["mrp"], 2)
                cgst, sgst, total = calculate_gst(taxable, gst_rate)
                st.session_state.cart.append({
                    "product_id":    product["id"],
                    "product_name":  product["name"],
                    "hsn_code":      product["hsn_code"],
                    "gst_rate":      gst_rate,
                    "quantity":      1.0,
                    "unit":          product["unit"],
                    "unit_price":    product["mrp"],
                    "taxable_value": taxable,
                    "cgst_amount":   cgst,
                    "sgst_amount":   sgst,
                    "line_total":    total,
                })
                st.toast(f"✅  {product['name']}  added  ₹{product['mrp']:.2f}", icon="🛒")
            st.rerun()

    # ── Manual fallback (hidden inside expander) ───────────────────────────
    with st.expander("⌨️  Add item manually (product has no QR code)"):
        all_prods = conn.execute(
            "SELECT id,name,hsn_code,unit,quantity,mrp,gst_rate FROM inventory ORDER BY name"
        ).fetchall()
        prod_map  = {p["name"]: dict(p) for p in all_prods}
        sel_name  = st.selectbox("Product", ["-- select --"] + list(prod_map.keys()))
        man_qty   = st.number_input("Qty", min_value=0.1, step=0.5, value=1.0)
        if st.button("➕ Add", key="manual_add"):
            if sel_name != "-- select --":
                prod     = prod_map[sel_name]
                gst_rate = prod.get("gst_rate") or 0.05
                if man_qty > prod["quantity"]:
                    st.error(f"❌  Only {prod['quantity']} {prod['unit']} available!")
                else:
                    taxable = round(prod["mrp"] * man_qty, 2)
                    cgst, sgst, total = calculate_gst(taxable, gst_rate)
                    existing = next((i for i,c in enumerate(st.session_state.cart)
                                     if c["product_id"] == prod["id"]), None)
                    if existing is not None:
                        st.session_state.cart[existing]["quantity"] += man_qty
                        c_ = st.session_state.cart[existing]
                        new_tax = round(c_["unit_price"] * c_["quantity"], 2)
                        (c_["taxable_value"], c_["cgst_amount"],
                         c_["sgst_amount"],   c_["line_total"]) = (
                            new_tax, *calculate_gst(new_tax, gst_rate))
                    else:
                        st.session_state.cart.append({
                            "product_id":    prod["id"],
                            "product_name":  prod["name"],
                            "hsn_code":      prod["hsn_code"],
                            "gst_rate":      gst_rate,
                            "quantity":      man_qty,
                            "unit":          prod["unit"],
                            "unit_price":    prod["mrp"],
                            "taxable_value": taxable,
                            "cgst_amount":   cgst,
                            "sgst_amount":   sgst,
                            "line_total":    total,
                        })
                    st.rerun()

    # ══════════════════════════════════════════════════════════════════════
    # STEP 2 — CART  (live, simple table)
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("""
    <div style="background:#2e7d32;border-radius:12px;padding:14px 22px;margin:18px 0 12px 0;">
        <h2 style="color:white;margin:0;font-size:1.3rem;">🛒 Step 2 — Items in Cart</h2>
    </div>
    """, unsafe_allow_html=True)

    if st.session_state.cart:

        # Simple clean cart table
        cart_display = []
        for item in st.session_state.cart:
            cart_display.append({
                "Product":   item["product_name"],
                "Qty":       item["quantity"],
                "Unit":      item["unit"],
                "MRP (₹)":   f"{item['unit_price']:.2f}",
                "Sale Rate": f"{item.get('sale_price', item['unit_price']):.2f}",
                "Total (₹)": f"{item['line_total']:.2f}",
            })
        st.dataframe(
            pd.DataFrame(cart_display),
            use_container_width=True,
            hide_index=True
        )

        # Remove buttons — one per item
        st.caption("Tap ✖ to remove an item:")
        cols = st.columns(min(len(st.session_state.cart), 4))
        for i, item in enumerate(st.session_state.cart):
            if cols[i % 4].button(
                f"✖ {item['product_name'][:18]}",
                key=f"del_{i}",
                use_container_width=True
            ):
                st.session_state.cart.pop(i)
                st.rerun()

        # ── Subtotal before discount ───────────────────────────────────────
        sub_taxable = sum(c["taxable_value"] for c in st.session_state.cart)
        sub_cgst    = sum(c["cgst_amount"]   for c in st.session_state.cart)
        sub_sgst    = sum(c["sgst_amount"]   for c in st.session_state.cart)
        sub_total   = sum(c["line_total"]    for c in st.session_state.cart)

        # ══════════════════════════════════════════════════════════════════
        # DISCOUNT SECTION — optional, two modes
        # ══════════════════════════════════════════════════════════════════
        st.markdown("""
        <div style="background:#f3e5f5;border-left:5px solid #7b1fa2;
                    border-radius:8px;padding:12px 18px;margin:12px 0 8px 0;">
            <b>🏷️ Discount (Optional)</b> — Leave at 0 if no discount
        </div>
        """, unsafe_allow_html=True)

        disc_col1, disc_col2 = st.columns(2)
        disc_type = disc_col1.radio(
            "Discount Type",
            ["₹ Fixed Amount", "% Percentage"],
            horizontal=True,
            label_visibility="collapsed"
        )
        disc_val = disc_col2.number_input(
            "Discount value",
            min_value=0.0,
            max_value=sub_total if disc_type == "₹ Fixed Amount" else 100.0,
            step=0.5,
            value=0.0,
            label_visibility="collapsed",
            format="%.2f"
        )

        # Calculate discount amount
        if disc_type == "% Percentage":
            discount_amt = round(sub_total * disc_val / 100, 2)
        else:
            discount_amt = round(min(disc_val, sub_total), 2)

        # Final amounts after discount
        grand_total   = round(sub_total   - discount_amt, 2)
        # Proportionally reduce taxable/cgst/sgst for record keeping
        ratio         = grand_total / sub_total if sub_total > 0 else 1
        grand_taxable = round(sub_taxable * ratio, 2)
        grand_cgst    = round(sub_cgst    * ratio, 2)
        grand_sgst    = round(sub_sgst    * ratio, 2)

        # Show totals box
        if discount_amt > 0:
            st.markdown(f"""
            <div style="background:#fff8e1;border:2px solid #f9a825;border-radius:12px;
                        padding:16px 24px;margin:12px 0;text-align:center;">
                <p style="margin:0;color:#555;font-size:0.9rem;">
                    Subtotal &nbsp;₹{sub_total:.2f}
                    &nbsp; — &nbsp;
                    <span style="color:#c62828;font-weight:bold;">
                        Discount&nbsp;
                        {"(" + str(disc_val) + "%)" if disc_type == "% Percentage" else ""}
                        &nbsp;₹{discount_amt:.2f}
                    </span>
                </p>
                <h1 style="margin:6px 0 0 0;color:#1a3c5e;font-size:2.2rem;">
                    TOTAL &nbsp; ₹ {grand_total:.2f}
                </h1>
                <p style="margin:4px 0 0 0;color:#777;font-size:0.8rem;">
                    Taxable ₹{grand_taxable:.2f} + CGST ₹{grand_cgst:.2f} + SGST ₹{grand_sgst:.2f}
                </p>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div style="background:#fff8e1;border:2px solid #f9a825;border-radius:12px;
                        padding:16px 24px;margin:12px 0;text-align:center;">
                <p style="margin:0;color:#555;font-size:0.9rem;">
                    Taxable ₹{grand_taxable:.2f} &nbsp;+&nbsp;
                    CGST ₹{grand_cgst:.2f} &nbsp;+&nbsp;
                    SGST ₹{grand_sgst:.2f}
                </p>
                <h1 style="margin:4px 0 0 0;color:#1a3c5e;font-size:2.2rem;">
                    TOTAL &nbsp; ₹ {grand_total:.2f}
                </h1>
            </div>
            """, unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════════════
        # STEP 3 — ONE BIG BUTTON
        # ══════════════════════════════════════════════════════════════════
        st.markdown("""
        <div style="background:#e8f5e9;border-radius:12px;padding:14px 22px;margin-bottom:12px;">
            <h2 style="color:#1b5e20;margin:0;font-size:1.3rem;">🖨️ Step 3 — Generate Bill</h2>
            <p style="color:#555;margin:4px 0 0 0;font-size:0.85rem;">
                Optional: enter customer name below. Then press the green button.
            </p>
        </div>
        """, unsafe_allow_html=True)

        cust_name_input = st.text_input(
            "Customer Name (optional)",
            value=st.session_state.bill_cust_name,
            placeholder="Leave empty for Cash Sale"
        )
        if cust_name_input:
            st.session_state.bill_cust_name = cust_name_input

        # THE BIG GREEN BUTTON
        if st.button(
            "✅   GENERATE BILL & DOWNLOAD PDF",
            type="primary",
            use_container_width=True
        ):
            inv_no   = generate_invoice_no()
            inv_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            c_name   = st.session_state.bill_cust_name or "Cash Customer"

            conn.execute("""
                INSERT INTO invoices
                    (invoice_no,customer_name,customer_phone,customer_gstin,
                     invoice_date,taxable_value,cgst_amount,sgst_amount,total_amount)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (inv_no, c_name, "", "",
                  inv_date, grand_taxable, grand_cgst, grand_sgst, grand_total))

            for item in st.session_state.cart:
                conn.execute("""
                    INSERT INTO invoice_items
                        (invoice_no,product_id,product_name,hsn_code,
                         quantity,unit,unit_price,gst_rate,
                         taxable_value,cgst_amount,sgst_amount,line_total)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """, (inv_no, item["product_id"], item["product_name"],
                      item["hsn_code"], item["quantity"], item["unit"],
                      item["unit_price"], item.get("gst_rate", 0.05),
                      item["taxable_value"], item["cgst_amount"],
                      item["sgst_amount"], item["line_total"]))
                conn.execute(
                    "UPDATE inventory SET quantity=quantity-? WHERE id=?",
                    (item["quantity"], item["product_id"])
                )
                rem = conn.execute(
                    "SELECT quantity,name FROM inventory WHERE id=?",
                    (item["product_id"],)
                ).fetchone()
                if rem and rem["quantity"] < LOW_STOCK_ALERT:
                    st.warning(f"⚠️  Low stock: {rem['name']} → {rem['quantity']} left")

            conn.commit()

            cart_snapshot = list(st.session_state.cart)
            pdf_bytes = generate_pdf_invoice({
                "invoice_no":    inv_no,
                "invoice_date":  inv_date,
                "customer_name": c_name,
                "customer_phone":"",
                "customer_gstin":"",
                "taxable_value": grand_taxable,
                "cgst_amount":   grand_cgst,
                "sgst_amount":   grand_sgst,
                "total_amount":  grand_total,
                "discount_amt":  discount_amt,
                "sub_total":     sub_total,
            }, cart_snapshot)

            st.session_state.cart           = []
            st.session_state.bill_cust_name = ""

            st.success(f"✅  Invoice {inv_no} saved!  Total ₹{grand_total:.2f}"
                       + (f"  (Discount ₹{discount_amt:.2f})" if discount_amt > 0 else ""))
            st.download_button(
                label="⬇️   DOWNLOAD PDF BILL",
                data=pdf_bytes,
                file_name=f"Bill_{inv_no}.pdf",
                mime="application/pdf",
                use_container_width=True
            )

        # Clear cart button — smaller, less prominent
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️  Clear cart and start over", use_container_width=False):
            st.session_state.cart           = []
            st.session_state.bill_cust_name = ""
            st.rerun()

    else:
        # Empty state — clear instructions
        st.markdown("""
        <div style="text-align:center;padding:50px 20px;background:#f9f9f9;
                    border-radius:12px;border:2px dashed #ccc;margin-top:10px;">
            <div style="font-size:3rem;">📱</div>
            <h3 style="color:#1a3c5e;">No items yet</h3>
            <p style="color:#777;font-size:1rem;">
                Click the scan box at the top<br>
                and scan a product QR code to begin.
            </p>
        </div>
        """, unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════════════════════════
# PAGE — REPORTS
# ═════════════════════════════════════════════════════════════════════════════
elif page == "📊 Reports":
    st.subheader("📊 CA-Ready Sales Reports — GSTR-1")

    today      = datetime.date.today()
    week_start = today - datetime.timedelta(days=today.weekday())

    daily_total   = conn.execute(
        "SELECT COALESCE(SUM(total_amount),0) FROM invoices WHERE DATE(invoice_date)=?",
        (today.isoformat(),)).fetchone()[0]
    weekly_total  = conn.execute(
        "SELECT COALESCE(SUM(total_amount),0) FROM invoices WHERE DATE(invoice_date) BETWEEN ? AND ?",
        (week_start.isoformat(), (week_start + datetime.timedelta(days=6)).isoformat())).fetchone()[0]
    monthly_total = conn.execute(
        "SELECT COALESCE(SUM(total_amount),0) FROM invoices WHERE strftime('%Y-%m',invoice_date)=?",
        (today.strftime("%Y-%m"),)).fetchone()[0]
    inv_count     = conn.execute("SELECT COUNT(*) FROM invoices").fetchone()[0]

    mc1,mc2,mc3,mc4 = st.columns(4)
    mc1.metric("Today",          f"₹ {daily_total:,.2f}")
    mc2.metric("This Week",      f"₹ {weekly_total:,.2f}")
    mc3.metric("This Month",     f"₹ {monthly_total:,.2f}")
    mc4.metric("Total Invoices", inv_count)

    st.markdown("---")
    if st.button("📥 Generate GSTR-1 Excel Report", type="primary"):
        with st.spinner("Generating..."):
            excel_bytes = generate_excel_report()
        st.download_button(
            "⬇️ Download Excel (Daily | Weekly | Monthly)",
            data=excel_bytes,
            file_name=f"SLV_Sales_{today.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success("✅ Share this with your CA for GSTR-1 filing.")

    st.markdown("---")
    st.markdown("#### 🗂️ Invoice History")
    search_inv = st.text_input("🔍 Search by Invoice No or Customer")
    import pandas as pd
    q  = """SELECT invoice_no,invoice_date,customer_name,
                   taxable_value,cgst_amount,sgst_amount,total_amount
            FROM invoices"""
    p  = ()
    if search_inv:
        q += " WHERE invoice_no LIKE ? OR customer_name LIKE ?"
        p  = (f"%{search_inv}%", f"%{search_inv}%")
    q += " ORDER BY id DESC LIMIT 50"
    rows = conn.execute(q, p).fetchall()
    if rows:
        df = pd.DataFrame([dict(r) for r in rows])
        df.columns = ["Invoice No","Date","Customer","Taxable(₹)","CGST(₹)","SGST(₹)","Total(₹)"]
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.info("No invoices found.")

# ═════════════════════════════════════════════════════════════════════════════
# PAGE — SEARCH PRODUCT
# ═════════════════════════════════════════════════════════════════════════════
elif page == "🔍 Search Product":
    st.subheader("🔍 Product Location Search")
    search_query = st.text_input(
        "Search by product name or QR code...",
        placeholder="e.g. Urea, DAP, Zinc, or type/scan a barcode"
    )

    if search_query.strip():
        results = conn.execute("""
            SELECT name,qr_code,hsn_code,section,row_no,slot,quantity,unit,mrp,gst_rate
            FROM inventory
            WHERE name LIKE ? OR qr_code LIKE ?
            ORDER BY section,row_no,slot
        """, (f"%{search_query}%", f"%{search_query}%")).fetchall()

        if results:
            st.markdown(f"**{len(results)} result(s) found:**")
            for prod in results:
                sc      = "🔴" if prod["quantity"] < LOW_STOCK_ALERT else "🟢"
                gst_pct = (prod["gst_rate"] or 0) * 100
                st.markdown(f"""
                <div style="background:#f8fffe;border:1px solid #a5d6a7;
                            border-left:5px solid #2e7d32;padding:12px 16px;
                            border-radius:8px;margin-bottom:10px;">
                    <h4 style="margin:0;color:#1a3c5e;">🌱 {prod['name']}</h4>
                    <p style="margin:4px 0;color:#555;">
                        HSN: <b>{prod['hsn_code']}</b> &nbsp;|&nbsp;
                        QR Code: <b>{prod['qr_code'] or '—'}</b> &nbsp;|&nbsp;
                        GST: <b>{gst_pct:.0f}%</b>
                    </p>
                    <p style="margin:4px 0;">
                        📍 <b>Location:</b>
                        {prod['section']} &nbsp;›&nbsp; {prod['row_no']} &nbsp;›&nbsp;
                        <b>Slot {prod['slot']}</b>
                    </p>
                    <p style="margin:4px 0;">
                        {sc} <b>Stock:</b> {prod['quantity']} {prod['unit']}
                        &nbsp;&nbsp;|&nbsp;&nbsp;
                        💰 <b>MRP:</b> ₹ {prod['mrp']:.2f} / {prod['unit']}
                    </p>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.warning(f"No products found for **'{search_query}'**.")
    else:
        st.info("Start typing a product name or QR code to search.")
        st.markdown("---")
        st.markdown("#### 🗺️ Shop Layout Overview")
        import pandas as pd
        sections = conn.execute(
            "SELECT section,COUNT(*) as cnt FROM inventory GROUP BY section ORDER BY section"
        ).fetchall()
        for sec in sections:
            prods = conn.execute(
                """SELECT name,qr_code,row_no,slot,quantity,unit
                   FROM inventory WHERE section=? ORDER BY row_no,slot""",
                (sec["section"],)
            ).fetchall()
            with st.expander(f"📦 {sec['section']} ({sec['cnt']} products)"):
                for p in prods:
                    badge   = "🔴 LOW" if p["quantity"] < LOW_STOCK_ALERT else "🟢"
                    barcode = f"| 📱 QR: `{p['qr_code']}`" if p["qr_code"] else ""
                    st.markdown(
                        f"• **{p['name']}** — {p['row_no']}, Slot {p['slot']} "
                        f"| Stock: {p['quantity']} {p['unit']} {badge} {barcode}"
                    )

elif page in dict(MONTHS):
    page_month(page)

elif page == "📋 GST Registers":
    gst_ss_init()
    st.session_state["bills"]          = st.session_state.gst_bills
    st.session_state["sales_override"] = st.session_state.gst_overrides
    st.session_state["suppliers"]      = st.session_state.gst_suppliers
    st.markdown("## 📋 GST Registers")
    sub_pages = [("dashboard","🏠 Overview"),("11-2025","Nov 2025"),
                 ("12-2025","Dec 2025"),("01-2026","Jan 2026"),
                 ("02-2026","Feb 2026"),("03-2026","Mar 2026"),
                 ("04-2026","Apr 2026")]
    btn_cols = st.columns(len(sub_pages))
    for i,(mk,lbl) in enumerate(sub_pages):
        active = st.session_state.gst_sub_page == mk
        if btn_cols[i].button(lbl, key=f"gstsub_{mk}", use_container_width=True,
                              type="primary" if active else "secondary"):
            st.session_state.gst_sub_page = mk
            st.rerun()
    st.markdown("---")
    _sub = st.session_state.gst_sub_page
    if _sub == "dashboard":
        page_dashboard()
    elif _sub in dict(MONTHS):
        page_month(_sub)
    else:
        page_dashboard()
    with st.sidebar:
        st.markdown("---")
        st.markdown("**Gemini API Key** (AI bill scan)")
        _gkey = st.text_input("", value=st.session_state.gemini_key,
                              type="password", placeholder="AIza…",
                              label_visibility="collapsed", key="gemini_key_sb")
        if _gkey:
            st.session_state.gemini_key = _gkey
        st.caption("✓ Key set" if st.session_state.gemini_key else "Get free key: aistudio.google.com")

conn.close()
